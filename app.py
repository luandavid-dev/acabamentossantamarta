# app.py
import pytz
import os
import secrets
try:
    from dotenv import load_dotenv
    # Garante que o .env seja buscado no diretório do script
    dotenv_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env')
    if os.path.exists(dotenv_path):
        load_dotenv(dotenv_path)
    else:
        load_dotenv() # Fallback para comportamento padrão
except ImportError:
    pass
import hmac
import sqlite3
try:
    import pyodbc
except ImportError:
    pyodbc = None
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, jsonify, abort
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import json
from datetime import datetime
from datetime import datetime, timedelta

def parse_date_br(data_str, default=None):
    if not data_str:
        return default
    try:
        return datetime.strptime(data_str, "%d/%m/%Y")
    except ValueError:
        return default

# --- CONFIG ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "banco.db")
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
ALLOWED_EXT = {"pdf","doc","docx","xls","xlsx","jpg","jpeg","png","txt","zip","csv"}
MAX_CONTENT_LENGTH = 50 * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app = Flask(__name__, template_folder="templates", static_folder="static")

def _get_or_create_secret_key():
    """Obtém SECRET_KEY do ambiente ou cria uma chave local persistente fora do código-fonte."""
    env_secret = os.environ.get("SECRET_KEY")
    if env_secret and len(env_secret) >= 32:
        return env_secret

    instance_dir = os.path.join(BASE_DIR, "instance")
    os.makedirs(instance_dir, exist_ok=True)
    secret_file = os.path.join(instance_dir, "secret_key")

    if os.path.exists(secret_file):
        with open(secret_file, "r", encoding="utf-8") as fh:
            value = fh.read().strip()
            if value:
                return value

    value = secrets.token_urlsafe(64)
    with open(secret_file, "w", encoding="utf-8") as fh:
        fh.write(value)
    try:
        os.chmod(secret_file, 0o600)
    except OSError:
        pass
    return value

app.secret_key = _get_or_create_secret_key()
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = MAX_CONTENT_LENGTH
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "0").lower() in ("1", "true", "yes")
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=int(os.environ.get("SESSION_HOURS", "8")))
app.config["JSON_SORT_KEYS"] = False

LOGIN_ATTEMPTS = {}
LOGIN_MAX_ATTEMPTS = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "5"))
LOGIN_WINDOW_SECONDS = int(os.environ.get("LOGIN_WINDOW_SECONDS", "900"))

def _client_ip():
    return (request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip())

def _is_login_blocked(usuario):
    key = (_client_ip(), (usuario or "").lower())
    now_ts = datetime.now(pytz.UTC).timestamp()
    attempts = [ts for ts in LOGIN_ATTEMPTS.get(key, []) if now_ts - ts < LOGIN_WINDOW_SECONDS]
    LOGIN_ATTEMPTS[key] = attempts
    return len(attempts) >= LOGIN_MAX_ATTEMPTS

def _register_login_failure(usuario):
    key = (_client_ip(), (usuario or "").lower())
    LOGIN_ATTEMPTS.setdefault(key, []).append(datetime.now(pytz.UTC).timestamp())

def _clear_login_failures(usuario):
    LOGIN_ATTEMPTS.pop((_client_ip(), (usuario or "").lower()), None)

def allowed_file(filename, allowed=None):
    allowed = allowed or ALLOWED_EXT
    return bool(filename and "." in filename and filename.rsplit(".", 1)[1].lower() in allowed)

def safe_unique_filename(filename, prefix=""):
    safe = secure_filename(filename or "")
    if not safe or "." not in safe:
        raise ValueError("Nome de arquivo inválido.")
    ext = safe.rsplit(".", 1)[1].lower()
    base = safe.rsplit(".", 1)[0][:80] or "arquivo"
    token = secrets.token_hex(8)
    ts = datetime.now(pytz.UTC).strftime("%Y%m%d%H%M%S")
    prefix = f"{secure_filename(prefix)}_" if prefix else ""
    return f"{prefix}{ts}_{token}_{base}.{ext}"

# OPTIONAL: path to local xlsx to try import on startup. Desativado por padrão em produção.
LOCAL_XLSX_PATH = os.environ.get("IMPORT_XLSX_PATH", "").strip()


# --- DB HELPERS ---
def get_db_connection():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db_connection()
    c = conn.cursor()

    # existing tables...
    c.execute("""CREATE TABLE IF NOT EXISTS filiais (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL,
        endereco TEXT,
        cidade TEXT,
        estado TEXT,
        telefone TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT NOT NULL,
        usuario TEXT UNIQUE,
        email TEXT UNIQUE,
        senha TEXT NOT NULL,
        tipo TEXT NOT NULL DEFAULT 'usuario',
        filial_id INTEGER
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS chamados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT NOT NULL,
        descricao TEXT,
        status TEXT NOT NULL DEFAULT 'Aberto',
        prioridade TEXT DEFAULT 'Normal',
        criado_por INTEGER,
        tecnico_id INTEGER,
        filial_id INTEGER,
        data_abertura TEXT,
        data_resolucao TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS anexos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chamado_id INTEGER,
        filename TEXT,
        filepath TEXT,
        uploaded_by INTEGER,
        uploaded_at TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS mensagens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        chamado_id INTEGER NOT NULL,
        autor_id INTEGER NOT NULL,
        autor_nome TEXT NOT NULL,
        autor_tipo TEXT NOT NULL,
        texto TEXT,
        anexo_filename TEXT,
        anexo_path TEXT,
        data TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS notificacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        mensagem_id INTEGER,
        chamado_id INTEGER,
        texto TEXT,
        autor_id INTEGER,
        autor_nome TEXT,
        lida INTEGER DEFAULT 0,
        criado_em TEXT
    )""")

    # NOVA TABELA: produtos (codpro + descricao_longa) - codpro único
    c.execute("""
        CREATE TABLE IF NOT EXISTS produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codpro TEXT NOT NULL,
            descricao_longa TEXT NOT NULL,
            UNIQUE(codpro)
        )
    """)

    conn.commit()
    conn.close()


def criar_admin_padrao():
    """Cria administrador inicial somente quando a senha vier do ambiente.

    Em produção, defina ADMIN_INITIAL_PASSWORD antes da primeira execução.
    O projeto não cria mais usuário administrativo com senha fraca conhecida.
    """
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT id FROM usuarios WHERE tipo = ? LIMIT 1", ("admin",))
    existe_admin = c.fetchone()
    senha_inicial = os.environ.get("ADMIN_INITIAL_PASSWORD", "").strip()
    if not existe_admin and senha_inicial:
        usuario_admin = os.environ.get("ADMIN_USERNAME", "admin").strip() or "admin"
        email_admin = os.environ.get("ADMIN_EMAIL", usuario_admin).strip() or usuario_admin
        senha_hash = generate_password_hash(senha_inicial)
        c.execute("INSERT INTO usuarios (nome, usuario, email, senha, tipo) VALUES (?,?,?,?,?)",
                  ("Administrador", usuario_admin, email_admin, senha_hash, "admin"))
        conn.commit()
    elif not existe_admin:
        print("[segurança] Nenhum administrador encontrado. Defina ADMIN_INITIAL_PASSWORD para criar o primeiro admin.")
    conn.close()

def extrair_classificacao(titulo: str):
    if not titulo:
        return "Outros"

    t = titulo.lower()

    if "quebras na entrega" in t:
        return "Quebras na Entrega"
    if "defeito de fábrica" in t or "defeito de fabrica" in t:
        return "Defeito de Fábrica"
    if "avarias de pátio" in t or "avarias de patio" in t:
        return "Avarias de Pátio"

    return "Outros"

# -------------------------------------------------------
# 4. FUNÇÃO PARA CRIAR NOTIFICAÇÃO  (COLOQUE AQUI!!!)
# -------------------------------------------------------
def criar_notificacao(chamado_id, texto, autor_id, autor_nome):
    conn = get_db_connection()
    now = datetime.now(pytz.UTC).isoformat()

    conn.execute("""
        INSERT INTO notificacoes (chamado_id, texto, autor_id, autor_nome, lida, criado_em)
        VALUES (?, ?, ?, ?, 0, ?)
    """, (chamado_id, texto, autor_id, autor_nome, now))

    conn.commit()
    conn.close()

# initialize DB + admin
init_db()
criar_admin_padrao()


# --- AUTENTICAÇÃO ---
def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapped


@app.context_processor
def inject_now():
    return {"now": lambda: datetime.now(pytz.UTC), "year": datetime.now(pytz.UTC).year}


# --- ROUTES (login / uploads) ---
@app.route("/uploads/<path:filename>")
@login_required
def download_upload(filename):
    if os.path.basename(filename) != filename or ".." in filename:
        abort(400)
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename, as_attachment=True)


@app.route("/", methods=["GET","POST"])
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        usuario = request.form.get("usuario","").strip()
        senha = request.form.get("senha","").strip()

        if _is_login_blocked(usuario):
            flash("Muitas tentativas inválidas. Aguarde alguns minutos e tente novamente.", "danger")
            return redirect(url_for("login"))

        conn = get_db_connection()
        user = conn.execute("SELECT * FROM usuarios WHERE usuario=? OR email=?", (usuario, usuario)).fetchone()
        conn.close()

        if not user or not check_password_hash(user["senha"], senha):
            _register_login_failure(usuario)
            flash("Credenciais inválidas.", "danger")
            return redirect(url_for("login"))

        _clear_login_failures(usuario)
        session.clear()
        session.permanent = True
        session["user_id"] = user["id"]
        session["nome"] = user["nome"]
        session["tipo"] = user["tipo"]

        return redirect(url_for("dashboard"))

    return render_template("login.html", year=datetime.now(pytz.UTC).year)


@app.route("/logout")
def logout():
    session.clear()
    flash("Você saiu do sistema.", "info")
    return redirect(url_for("login"))

import re
from markupsafe import Markup
# certifique-se de já ter get_db_connection definido no app.py

def _extrair_codigo_do_titulo(titulo: str) -> str:
    if not titulo:
        return ""
    # código antes do primeiro " - "
    parts = titulo.split(" - ", 1)
    return parts[0].strip() if parts else titulo.strip()

def _extrair_quantidade_do_titulo(titulo: str) -> str:
    if not titulo:
        return ""
    # tenta capturar depois do último " - "
    if " - " in titulo:
        last = titulo.rsplit(" - ", 1)[1].strip()
        return last
    return ""

@app.route("/dashboard")
@login_required
def dashboard():

    from flask import session, redirect, url_for, request, render_template
    import re

    if session.get("tipo") != "admin":

        return redirect(
            url_for("listar_chamados")
        )

    conn = get_db_connection()

    conn.row_factory = sqlite3.Row

    filtros = {
        "status": request.args.get("status", "").strip(),
        "status_ajuste": request.args.get("status_ajuste", "").strip(),
        "produto": request.args.get("produto", "").strip(),
        "usuario": request.args.get("usuario", "").strip(),
        "inicio": request.args.get("inicio", "").strip(),
        "fim": request.args.get("fim", "").strip(),
    }

    def numero(valor):

        try:
            return float(valor or 0)
        except:
            return 0

    def formatar_brl(valor):

        try:
            texto = f"{float(valor or 0):,.2f}"

            return (
                texto
                .replace(",", "X")
                .replace(".", ",")
                .replace("X", ".")
            )

        except:
            return "0,00"

    def obter_colunas(tabela):

        try:

            rows = conn.execute(f"""
                PRAGMA table_info({tabela})
            """).fetchall()

            return {
                row["name"]
                for row in rows
            }

        except:
            return set()

    def normalizar_status_ajuste(status):

        status = (
            str(status or "")
            .strip()
            .upper()
        )

        mapa = {
            "": "ABERTO",
            "ABERTO": "ABERTO",
            "EM ANALISE": "EM ANALISE",
            "EM ANÁLISE": "EM ANALISE",
            "EM ANDAMENTO": "EM ANALISE",
            "FINALIZADO": "FINALIZADO",
            "CONCLUIDO": "FINALIZADO",
            "CONCLUÍDO": "FINALIZADO",
        }

        return mapa.get(
            status,
            status or "ABERTO"
        )

    def extrair_quantidade_unidade(titulo):

        titulo = str(titulo or "").strip()

        match = re.search(
            r"(\d+[.,]?\d*)\s*(UN|m²|M2|P[Cc])?$",
            titulo
        )

        if match:

            try:

                return (
                    float(match.group(1).replace(",", ".")),
                    match.group(2) or "UN"
                )

            except:
                return 1, "UN"

        return 1, "UN"

    # =====================================================
    # REQUERIMENTOS / CHAMADOS
    # =====================================================

    chamados_where = []
    chamados_params = []

    if filtros["status"]:

        chamados_where.append("c.status = ?")
        chamados_params.append(filtros["status"])

    if filtros["produto"]:

        chamados_where.append("c.titulo LIKE ?")
        chamados_params.append(f"%{filtros['produto']}%")

    if filtros["inicio"]:

        chamados_where.append("date(c.data_abertura) >= date(?)")
        chamados_params.append(filtros["inicio"])

    if filtros["fim"]:

        chamados_where.append("date(c.data_abertura) <= date(?)")
        chamados_params.append(filtros["fim"])

    chamados_where_sql = (
        "WHERE " + " AND ".join(chamados_where)
        if chamados_where
        else ""
    )

    chamados_rows = conn.execute(f"""

        SELECT
            c.titulo AS produto,
            c.status,
            CASE
                WHEN c.prioridade IN (
                    'Quebras na Entrega',
                    'Defeito de Fábrica',
                    'Avarias de Pátio'
                ) THEN c.prioridade
                ELSE 'Outros'
            END AS classificacao

        FROM chamados c

        {chamados_where_sql}

    """, chamados_params).fetchall()

    total_chamados = 0

    status_counts = {
        "Aberto": 0,
        "Concluído": 0,
        "Em Andamento": 0
    }

    chamados_produto_classificacao = {}

    for row in chamados_rows:

        total_chamados += 1

        status_nome = row["status"] or "Sem Status"

        status_counts[status_nome] = (
            status_counts.get(status_nome, 0)
            + 1
        )

        titulo_limpo = re.sub(
            r"\s*-\s*\d+[.,]?\d*\s*(UN|m²|M2|P[Cc])?$",
            "",
            row["produto"] or ""
        ).strip()

        qtd, unidade = extrair_quantidade_unidade(
            row["produto"]
        )

        key = (
            titulo_limpo or "Sem produto",
            row["classificacao"] or "Outros",
            unidade
        )

        chamados_produto_classificacao[key] = (
            chamados_produto_classificacao.get(key, 0)
            + qtd
        )

    chamados_produto_classificacao_list = [

        {
            "produto": key[0],
            "classificacao": key[1],
            "total": round(valor, 2),
            "unidade": key[2],
        }

        for key, valor in chamados_produto_classificacao.items()

    ]

    chamados_produto_classificacao_list.sort(
        key=lambda item: item["total"],
        reverse=True
    )

    percentual_chamados_abertos = (
        round(
            (
                status_counts.get("Aberto", 0)
                / total_chamados
            ) * 100,
            1
        )
        if total_chamados
        else 0
    )

    # =====================================================
    # AJUSTES DE ESTOQUE
    # =====================================================

    furo_colunas = obter_colunas("furo_estoque")

    desc_partes = []

    if "descricao" in furo_colunas:
        desc_partes.append("NULLIF(f.descricao,'')")

    if "produto" in furo_colunas:
        desc_partes.append("NULLIF(f.produto,'')")

    descricao_expr = (
        "COALESCE("
        + ", ".join(desc_partes)
        + ", 'SEM DESCRIÇÃO')"
        if desc_partes
        else "'SEM DESCRIÇÃO'"
    )

    codpro_expr = (
        "COALESCE(NULLIF(f.codpro,''), '-')"
        if "codpro" in furo_colunas
        else "'-'"
    )

    status_expr = (
        "COALESCE(NULLIF(f.status,''), 'ABERTO')"
        if "status" in furo_colunas
        else "'ABERTO'"
    )

    filial_expr = (
        "COALESCE(NULLIF(f.filial,''), 'NÃO INFORMADA')"
        if "filial" in furo_colunas
        else "'NÃO INFORMADA'"
    )

    patio_expr = (
        "COALESCE(NULLIF(f.patio,''), 'NÃO INFORMADO')"
        if "patio" in furo_colunas
        else "'NÃO INFORMADO'"
    )

    joins_furos = []

    usuario_partes = []

    if "usuario_nome" in furo_colunas:
        usuario_partes.append("NULLIF(f.usuario_nome,'')")

    if "conferido_por" in furo_colunas:

        joins_furos.append("""
            LEFT JOIN usuarios uc
                ON uc.id = f.conferido_por
        """)

        usuario_partes.append("uc.nome")

    if "usuario_id" in furo_colunas:

        joins_furos.append("""
            LEFT JOIN usuarios ua
                ON ua.id = f.usuario_id
        """)

        usuario_partes.append("ua.nome")

    usuario_expr = (
        "COALESCE("
        + ", ".join(usuario_partes)
        + ", '-')"
        if usuario_partes
        else "'-'"
    )

    furos_where = []
    furos_params = []

    if filtros["produto"]:

        produto_condicoes = []

        if "codpro" in furo_colunas:

            produto_condicoes.append("f.codpro LIKE ?")
            furos_params.append(f"%{filtros['produto']}%")

        if "descricao" in furo_colunas:

            produto_condicoes.append("f.descricao LIKE ?")
            furos_params.append(f"%{filtros['produto']}%")

        if "produto" in furo_colunas:

            produto_condicoes.append("f.produto LIKE ?")
            furos_params.append(f"%{filtros['produto']}%")

        if produto_condicoes:

            furos_where.append(
                "(" + " OR ".join(produto_condicoes) + ")"
            )

    if filtros["usuario"]:

        usuario_condicoes = []

        if "usuario_nome" in furo_colunas:

            usuario_condicoes.append("f.usuario_nome LIKE ?")
            furos_params.append(f"%{filtros['usuario']}%")

        if "conferido_por" in furo_colunas:

            usuario_condicoes.append("uc.nome LIKE ?")
            furos_params.append(f"%{filtros['usuario']}%")

        if "usuario_id" in furo_colunas:

            usuario_condicoes.append("ua.nome LIKE ?")
            furos_params.append(f"%{filtros['usuario']}%")

        if usuario_condicoes:

            furos_where.append(
                "(" + " OR ".join(usuario_condicoes) + ")"
            )

    if filtros["status_ajuste"] and "status" in furo_colunas:

        furos_where.append("""
            COALESCE(NULLIF(f.status,''), 'ABERTO') = ?
        """)

        furos_params.append(
            normalizar_status_ajuste(
                filtros["status_ajuste"]
            )
        )

    if filtros["inicio"] and "data_furo" in furo_colunas:

        furos_where.append("date(f.data_furo) >= date(?)")
        furos_params.append(filtros["inicio"])

    if filtros["fim"] and "data_furo" in furo_colunas:

        furos_where.append("date(f.data_furo) <= date(?)")
        furos_params.append(filtros["fim"])

    furos_where_sql = (
        "WHERE " + " AND ".join(furos_where)
        if furos_where
        else ""
    )

    if "id" in furo_colunas:

        ordem_furos = "ORDER BY f.id DESC"

    elif "data_furo" in furo_colunas:

        ordem_furos = "ORDER BY f.data_furo DESC"

    else:

        ordem_furos = ""

    furos_rows = conn.execute(f"""

        SELECT
            f.*,
            {codpro_expr} AS codpro_calc,
            {descricao_expr} AS descricao_calc,
            {status_expr} AS status_calc,
            {filial_expr} AS filial_calc,
            {patio_expr} AS patio_calc,
            {usuario_expr} AS usuario_calc

        FROM furo_estoque f

        {" ".join(joins_furos)}

        {furos_where_sql}

        {ordem_furos}

    """, furos_params).fetchall()

    total_ajustes = 0
    total_corretos = 0
    total_divergentes = 0
    total_qtde_encontrada = 0

    ajuste_status_counts = {
        "ABERTO": 0,
        "EM ANALISE": 0,
        "FINALIZADO": 0
    }

    produtos_furos = {}
    filiais_furos = {}
    patios_furos = {}
    ajustes_recentes = []

    for row in furos_rows:

        item = dict(row)

        total_ajustes += 1

        qtde_encontrada = numero(
            item.get("qtde_encontrada")
        )

        qtde_sistema = numero(
            item.get("qtde_sistema")
        )

        dif = qtde_encontrada - qtde_sistema

        total_qtde_encontrada += qtde_encontrada

        if round(dif, 2) == 0:
            total_corretos += 1
        else:
            total_divergentes += 1

        status_ajuste = normalizar_status_ajuste(
            item.get("status_calc")
        )

        ajuste_status_counts[status_ajuste] = (
            ajuste_status_counts.get(status_ajuste, 0)
            + 1
        )

        codpro = item.get("codpro_calc") or "-"

        descricao = item.get("descricao_calc") or "SEM DESCRIÇÃO"

        produto_nome = (
            f"{codpro} — {descricao}"
            if codpro != "-"
            else descricao
        )

        produto_key = (
            codpro,
            descricao
        )

        if produto_key not in produtos_furos:

            produtos_furos[produto_key] = {
                "produto": produto_nome,
                "codpro": codpro,
                "descricao": descricao,
                "quantidade": 0,
                "total_qtde": 0,
                "dif_liquida": 0,
                "divergentes": 0
            }

        produtos_furos[produto_key]["quantidade"] += 1
        produtos_furos[produto_key]["total_qtde"] += qtde_encontrada
        produtos_furos[produto_key]["dif_liquida"] += dif

        if round(dif, 2) != 0:
            produtos_furos[produto_key]["divergentes"] += 1

        filial = item.get("filial_calc") or "NÃO INFORMADA"

        if filial not in filiais_furos:

            filiais_furos[filial] = {
                "filial": filial,
                "quantidade": 0,
                "divergentes": 0
            }

        filiais_furos[filial]["quantidade"] += 1

        if round(dif, 2) != 0:
            filiais_furos[filial]["divergentes"] += 1

        patio = item.get("patio_calc") or "NÃO INFORMADO"

        if patio not in patios_furos:

            patios_furos[patio] = {
                "patio": patio,
                "quantidade": 0,
                "divergentes": 0
            }

        patios_furos[patio]["quantidade"] += 1

        if round(dif, 2) != 0:
            patios_furos[patio]["divergentes"] += 1

        if len(ajustes_recentes) < 8:

            data_furo = item.get("data_furo")

            data_fmt = "-"

            if data_furo:

                try:
                    data_str = str(data_furo)
                    data_fmt = f"{data_str[8:10]}/{data_str[5:7]}/{data_str[0:4]}"
                except:
                    data_fmt = str(data_furo)

            ajustes_recentes.append({
                "id": item.get("id"),
                "codpro": codpro,
                "descricao": descricao,
                "filial": filial,
                "patio": patio,
                "status": status_ajuste,
                "dif": round(dif, 2),
                "usuario": item.get("usuario_calc") or "-",
                "data": data_fmt
            })

    top_produtos_furos = list(
        produtos_furos.values()
    )

    top_produtos_furos.sort(
        key=lambda item: (
            item["divergentes"],
            item["quantidade"],
            abs(item["dif_liquida"]),
            item["total_qtde"]
        ),
        reverse=True
    )

    top_produtos_furos_list = []

    for item in top_produtos_furos[:20]:

        top_produtos_furos_list.append({
            "produto": item["produto"],
            "codpro": item["codpro"],
            "descricao": item["descricao"],
            "quantidade": item["quantidade"],
            "total_qtde": round(item["total_qtde"], 2),
            "dif_liquida": round(item["dif_liquida"], 2),
            "divergentes": item["divergentes"]
        })

    ajustes_por_filial = list(
        filiais_furos.values()
    )

    ajustes_por_filial.sort(
        key=lambda item: (
            item["divergentes"],
            item["quantidade"]
        ),
        reverse=True
    )

    ajustes_por_patio = list(
        patios_furos.values()
    )

    ajustes_por_patio.sort(
        key=lambda item: (
            item["divergentes"],
            item["quantidade"]
        ),
        reverse=True
    )

    percentual_ajustes_divergentes = (
        round(
            (
                total_divergentes
                / total_ajustes
            ) * 100,
            1
        )
        if total_ajustes
        else 0
    )

    # =====================================================
    # COMPRAS
    # =====================================================

    compras_rows = conn.execute("""

        SELECT
            COALESCE(c.status_compra,'Não Informado') AS status,
            COUNT(a.id) AS quantidade,
            COALESCE(SUM(a.valor_total),0) AS total_valor

        FROM chamados c

        LEFT JOIN acompanhamento_compras a
            ON a.chamado_id = c.id

        GROUP BY COALESCE(c.status_compra,'Não Informado')

        ORDER BY total_valor DESC

    """).fetchall()

    compras_stats = [
        dict(row)
        for row in compras_rows
    ]

    compras_valor_total = sum(
        numero(item.get("total_valor"))
        for item in compras_stats
    )

    compras_quantidade_total = sum(
        int(item.get("quantidade") or 0)
        for item in compras_stats
    )

    charts = {
        "chamados_labels": [
            item["produto"]
            for item in chamados_produto_classificacao_list[:10]
        ],
        "chamados_values": [
            item["total"]
            for item in chamados_produto_classificacao_list[:10]
        ],
        "ajustes_produtos_labels": [
            item["produto"]
            for item in top_produtos_furos_list[:10]
        ],
        "ajustes_produtos_values": [
            item["total_qtde"]
            for item in top_produtos_furos_list[:10]
        ],
        "ajustes_status_labels": [
            "ABERTO",
            "EM ANALISE",
            "FINALIZADO"
        ],
        "ajustes_status_values": [
            ajuste_status_counts.get("ABERTO", 0),
            ajuste_status_counts.get("EM ANALISE", 0),
            ajuste_status_counts.get("FINALIZADO", 0)
        ],
        "compras_labels": [
            item["status"]
            for item in compras_stats
        ],
        "compras_values": [
            item["quantidade"]
            for item in compras_stats
        ]
    }

    usuarios_dashboard = conn.execute("""

        SELECT
            id,
            nome

        FROM usuarios

        ORDER BY nome

    """).fetchall()

    conn.close()

    # Obter contagem de filiais com tratamento de erro
    try:
        total_filiais = conn.execute("SELECT COUNT(*) FROM filiais").fetchone()[0]
    except Exception:
        total_filiais = 0

    # Obter furos do mês atual com tratamento de erro
    try:
        hoje_now = datetime.now()
        inicio_mes = hoje_now.replace(day=1).strftime("%Y-%m-%d")
        # Tenta data_registro ou data_furo dependendo da versão do banco
        try:
            furos_mes = conn.execute("SELECT COUNT(*) FROM furo_estoque WHERE date(data_registro) >= date(?)", (inicio_mes,)).fetchone()[0]
        except Exception:
            furos_mes = conn.execute("SELECT COUNT(*) FROM furo_estoque WHERE date(data_furo) >= date(?)", (inicio_mes,)).fetchone()[0]
    except Exception:
        furos_mes = 0

    # Preparar objeto stats para o template
    stats = {
        "chamados_abertos": status_counts.get("Aberto", 0),
        "entregas_hoje": len(compras_stats),
        "furos_mes": furos_mes,
        "total_filiais": total_filiais
    }

    # Ajustar objeto charts para os nomes esperados no template
    charts_template = {
        "filiais_labels": [item["produto"] for item in chamados_produto_classificacao_list[:10]],
        "filiais_values": [item["total"] for item in chamados_produto_classificacao_list[:10]],
        "status_labels": list(status_counts.keys()),
        "status_values": list(status_counts.values()),
        "furos_labels": [item["produto"] for item in top_produtos_furos_list[:10]],
        "furos_values": [item["total_qtde"] for item in top_produtos_furos_list[:10]],
        "compras_labels": [item["status"] for item in compras_stats],
        "compras_values": [item["quantidade"] for item in compras_stats]
    }

    return render_template(
        "dashboard.html",
        filtros=filtros,
        stats=stats,
        charts=charts_template,
        usuarios_dashboard=usuarios_dashboard
    )

# ----------------------------
# IMPORTAÇÃO: XLSX -> produtos
# ----------------------------
def import_products_from_xlsx(path):
    """
    Lê arquivo .xlsx e insere (codpro, DESCRICAO_LONGA) na tabela produtos.
    Procura colunas (case-insensitive): 'codpro' e 'DESCRICAO_LONGA' ou 'DESCRICAO LONGA'
    """
    if not os.path.exists(path):
        print(f"[import_products_from_xlsx] arquivo não encontrado: {path}")
        return {"ok": False, "error": "file not found"}

    try:
        from openpyxl import load_workbook
    except Exception as e:
        print("openpyxl não disponível. Instale com: pip install openpyxl")
        return {"ok": False, "error": "openpyxl missing"}

    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active

    # read header (first row)
    header = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        header.append(cell if cell is not None else "")

    # normalize headers -> map column index to header name
    normalized = [str(h).strip().lower() for h in header]

    # try to find columns
    try:
        idx_cod = None
        idx_desc = None
        for i, h in enumerate(normalized):
            if h in ("codpro", "cod_pro", "codigo", "codigo_produto", "codigoproduto"):
                idx_cod = i
            if h in ("descricao_longa", "descricaolonga", "descricao longa", "descricao_long", "descricao_long_a", "descricao_longab"):
                idx_desc = i
            # also allow DESCRICAO_LONGA uppercase original names
            if h == "descricao_curta" and idx_desc is None:
                # prefer descricao_longa but fallback later
                pass

        # If not found, try more fuzzy matches
        if idx_cod is None:
            for i, h in enumerate(normalized):
                if "cod" in h and ("pro" in h or "prod" in h or "codigo" in h):
                    idx_cod = i
                    break

        if idx_desc is None:
            for i, h in enumerate(normalized):
                if "descricao" in h and ("long" in h or "larga" in h or "completa" in h):
                    idx_desc = i
                    break

        if idx_cod is None or idx_desc is None:
            # try to guess from common table pattern (as user showed: Filial, codpro, DESCRICAO_CURTA, DESCRICAO_LONGA)
            for i, h in enumerate(normalized):
                if "filial" in h:
                    # common pattern, so codpro next column?
                    pass
            # final check: if 2nd column present and 4th column present assume those
            if idx_cod is None and len(normalized) >= 2:
                idx_cod = 1
            if idx_desc is None and len(normalized) >= 4:
                idx_desc = 3

        if idx_cod is None or idx_desc is None:
            return {"ok": False, "error": "columns not found", "cols": normalized}

        # iterate rows and collect
        rows_to_insert = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            cod = row[idx_cod] if idx_cod < len(row) else None
            desc = row[idx_desc] if idx_desc < len(row) else None

            if cod is None and desc is None:
                continue

            cod_s = str(cod).strip() if cod is not None else ""
            desc_s = str(desc).strip() if desc is not None else ""

            if not cod_s or not desc_s:
                continue

            rows_to_insert.append((cod_s, desc_s))

        # insert into DB (ignore duplicates by codpro)
        conn = get_db_connection()
        cur = conn.cursor()
        inserted = 0
        for cod_s, desc_s in rows_to_insert:
            try:
                cur.execute("INSERT OR IGNORE INTO produtos (codpro, descricao_longa) VALUES (?, ?)", (cod_s, desc_s))
                if cur.rowcount:
                    inserted += 1
            except Exception as e:
                print("Erro inserindo produto:", e)
        conn.commit()
        conn.close()

        print(f"[import_products_from_xlsx] import completo. linhas:{len(rows_to_insert)} inseridas:{inserted}")
        return {"ok": True, "rows": len(rows_to_insert), "inserted": inserted}

    except Exception as e:
        print("Erro ao importar xlsx:", e)
        return {"ok": False, "error": str(e)}


# Route para upload/importar via web (form)
@app.route("/importar_produtos", methods=["GET","POST"])
@login_required
def importar_produtos():
    # apenas admin pode importar
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        file = request.files.get("arquivo")
        if not file or not file.filename:
            flash("Nenhum arquivo enviado.", "warning")
            return redirect(url_for("importar_produtos"))

        if not allowed_file(file.filename, {"xlsx", "csv"}):
            flash("Formato não suportado. Envie .xlsx ou .csv", "warning")
            return redirect(url_for("importar_produtos"))
        filename = safe_unique_filename(file.filename, "produtos")
        save_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(save_path)

        # só aceita xlsx/csv
        ext = filename.rsplit(".", 1)[-1].lower()
        result = {"ok": False, "error": "extensão não suportada"}
        if ext in ("xlsx",):
            result = import_products_from_xlsx(save_path)
        elif ext in ("csv",):
            # quick csv import (semicolon or comma)
            try:
                import csv
                conn = get_db_connection()
                cur = conn.cursor()
                with open(save_path, "r", encoding="utf-8", errors="ignore") as f:
                    # detect delimiter: if semicolon present
                    sample = f.read(4096)
                    f.seek(0)
                    delim = ";" if ";" in sample and sample.count(";") > sample.count(",") else ","
                    reader = csv.DictReader(f, delimiter=delim)
                    inserted = 0
                    rows = 0
                    for row in reader:
                        rows += 1
                        cod = None
                        desc = None
                        for k, v in row.items():
                            key = k.strip().lower()
                            if key in ("codpro", "cod_pro", "codigo", "cod"):
                                cod = v
                            if key in ("descricao_longa", "descricaolonga", "descricao longa", "descricao_long"):
                                desc = v
                        if not cod:
                            # try common column names
                            cod = row.get("codpro") or row.get("CODPRO") or row.get(list(row.keys())[0])
                        if not desc and len(row.keys()) >= 2:
                            # try to pick last column
                            desc = row.get("DESCRICAO_LONGA") or row.get(list(row.keys())[-1])

                        if cod and desc:
                            cur.execute("INSERT OR IGNORE INTO produtos (codpro, descricao_longa) VALUES (?, ?)", (str(cod).strip(), str(desc).strip()))
                            if cur.rowcount:
                                inserted += 1
                conn.commit()
                conn.close()
                result = {"ok": True, "rows": rows, "inserted": inserted}
            except Exception as e:
                result = {"ok": False, "error": str(e)}
        else:
            flash("Formato não suportado. Envie .xlsx ou .csv", "warning")
            return redirect(url_for("importar_produtos"))

        if result.get("ok"):
            flash(f"Importação concluída. Inseridos: {result.get('inserted',0)}", "success")
        else:
            flash(f"Importação falhou: {result.get('error')}", "danger")

        return redirect(url_for("dashboard"))

    # GET -> mostra um formulário simples
    return render_template("importar_produtos.html")


# ----------------------------
# API: produtos (autocomplete)
# ----------------------------
@app.route("/api/produtos")
@login_required
def api_produtos():
    termo = request.args.get("q", "") or request.args.get("buscar", "") or ""
    termo = termo.strip()
    conn = get_db_connection()

    if termo:
        q = "%{}%".format(termo)
        rows = conn.execute("""
            SELECT codpro, descricao_longa
            FROM produtos
            WHERE codpro LIKE ? OR descricao_longa LIKE ?
            ORDER BY codpro
            LIMIT 50
        """, (q, q)).fetchall()
    else:
        rows = conn.execute("""
            SELECT codpro, descricao_longa
            FROM produtos
            ORDER BY codpro
            LIMIT 100
        """).fetchall()

    conn.close()

    # NOTE: return field names matching templates: codpro and DESCRICAO_LONGA (uppercase) are used in some templates,
    # but JSON keys are case-sensitive; many templates used prod.DESCRICAO_LONGA. To be safe, include both.
    result = []
    for r in rows:
        item = dict(r)
        result.append({
            "codpro": item.get("codpro"),
            "descricao_longa": item.get("descricao_longa"),
            "DESCRICAO_LONGA": item.get("descricao_longa")
        })

    return jsonify(result)


# ----------------------------
# RESTO DO SEU APP (chamados, usuarios, etc.)
# ----------------------------
# For concision I include the same handlers you had previously, slightly adapted to reference produto field where needed.
# (I kept your original implementations but ensured routes referenced below exist and templates use api/produtos as previously)
# ... (the following are your existing route handlers, repeated)

@app.route("/api/dashboard_stats")
@login_required
def api_dashboard_stats():
    conn = get_db_connection()

    filial_id = request.args.get("filial_id", type=int)
    classificacao = request.args.get("classificacao", type=str)
    inicio = request.args.get("inicio", type=str)
    fim = request.args.get("fim", type=str)

    where_clauses = []
    params = []

    if filial_id:
        where_clauses.append("ch.filial_id = ?")
        params.append(filial_id)

    if classificacao:
        where_clauses.append("ch.prioridade = ?")
        params.append(classificacao)

    if inicio and fim:
        where_clauses.append("date(ch.data_abertura) BETWEEN date(?) AND date(?)")
        params.extend([inicio, fim])
    elif inicio:
        where_clauses.append("date(ch.data_abertura) >= date(?)")
        params.append(inicio)
    elif fim:
        where_clauses.append("date(ch.data_abertura) <= date(?)")
        params.append(fim)

    where_sql = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

    sql_status = f"""
        SELECT ch.status AS key, COUNT(*) AS cnt
        FROM chamados ch
        {where_sql}
        GROUP BY ch.status
    """
    rows = conn.execute(sql_status, params).fetchall()
    by_status = {r["key"]: r["cnt"] for r in rows}

    sql_prio = f"""
        SELECT ch.prioridade AS key, COUNT(*) AS cnt
        FROM chamados ch
        {where_sql}
        GROUP BY ch.prioridade
    """
    rows = conn.execute(sql_prio, params).fetchall()
    by_prioridade = {r["key"]: r["cnt"] for r in rows}

    by_classificacao = by_prioridade.copy()

    sql_filial = f"""
        SELECT COALESCE(f.nome, 'Sem filial') AS filial, COUNT(*) AS cnt
        FROM chamados ch
        LEFT JOIN filiais f ON ch.filial_id = f.id
        {where_sql}
        GROUP BY ch.filial_id
        ORDER BY cnt DESC
    """
    rows = conn.execute(sql_filial, params).fetchall()
    by_filial = [{"filial": r["filial"], "cnt": r["cnt"]} for r in rows]

    conn.close()

    return jsonify({
        "by_status": by_status,
        "by_prioridade": by_prioridade,
        "by_classificacao": by_classificacao,
        "by_filial": by_filial
    })


@app.route("/api/filiais")
@login_required
def api_filiais():
    try:
        conn = get_db_connection()
        rows = conn.execute("SELECT id, nome FROM filiais ORDER BY nome").fetchall()
        conn.close()
        items = [{"id": r["id"], "nome": r["nome"]} for r in rows]
        return jsonify(items)
    except Exception:
        return jsonify([])

from datetime import datetime, timezone

def tempo_relativo(iso_time):
    if not iso_time:
        return ""
    dt = datetime.fromisoformat(iso_time).replace(tzinfo=timezone.utc)
    agora = datetime.now(pytz.UTC).replace(tzinfo=timezone.utc)
    diff = (agora - dt).total_seconds()

    if diff < 60:
        return "Agora mesmo"
    elif diff < 3600:
        return f"Há {int(diff/60)} min"
    elif diff < 86400:
        return f"Há {int(diff/3600)} h"
    else:
        return f"Há {int(diff/86400)} dias"

@app.route("/api/notificacoes")
def api_notificacoes():
    if 'user_id' not in session:
        return jsonify([])

    conn = get_db_connection()

    # Agora TODOS veem TUDO
    sql = """
        SELECT id, chamado_id, autor_nome, texto, lida, criado_em
        FROM notificacoes
        ORDER BY id DESC
        LIMIT 40
    """
    dados = conn.execute(sql).fetchall()
    conn.close()

    lista = []
    for n in dados:
        lista.append({
            "id": n["id"],
            "chamado_id": n["chamado_id"],
            "autor_nome": n["autor_nome"],
            "texto": n["texto"],
            "lida": bool(n["lida"]),
            "criado_em_formatado": tempo_relativo(n["criado_em"])
        })

    return jsonify(lista)

@app.route("/api/notificacoes/marcar/<int:notif_id>")
def api_notificacoes_marcar(notif_id):
    conn = get_db_connection()
    conn.execute("UPDATE notificacoes SET lida = 1 WHERE id = ?", (notif_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

@app.route("/api/notificacoes/limpar", methods=["POST"])
def api_notificacoes_limpar():
    if 'user_id' not in session:
        return jsonify({"erro": "não autorizado"}), 403

    uid = session['user_id']
    tipo = session.get("tipo")

    conn = get_db_connection()

    # Usuário comum só limpa suas notificações
    if tipo == "usuario":
        sql = """
            UPDATE notificacoes
            SET lida = 1
            WHERE id IN (
                SELECT n.id
                FROM notificacoes n
                JOIN chamados c ON c.id = n.chamado_id
                WHERE c.usuario_id = ?
            )
        """
        conn.execute(sql, (uid,))
    
    # Admin / Técnico limpa tudo
    else:
        conn.execute("UPDATE notificacoes SET lida = 1")

    conn.commit()
    conn.close()

    return jsonify({"status": "ok"})


# ---------- CHAMADOS (listar / novo / visualizar / mensagem / alterar status) ----------
@app.route("/chamados")
@login_required
def listar_chamados():
    conn = get_db_connection()
    tipo = session.get("tipo")
    user_id = session.get("user_id")

    if tipo == "usuario":
        rows = conn.execute("""
            SELECT ch.id, ch.titulo, ch.status, ch.prioridade, ch.data_abertura,
                   f.nome AS filial_nome, u.nome AS criador, t.nome AS tecnico
            FROM chamados ch
            LEFT JOIN usuarios u ON ch.criado_por = u.id
            LEFT JOIN usuarios t ON ch.tecnico_id = t.id
            LEFT JOIN filiais f ON ch.filial_id = f.id
            WHERE ch.criado_por = ?
            ORDER BY ch.id DESC
        """, (user_id,)).fetchall()

    else:
        rows = conn.execute("""
            SELECT ch.id, ch.titulo, ch.status, ch.prioridade, ch.data_abertura,
                   f.nome AS filial_nome, u.nome AS criador, t.nome AS tecnico
            FROM chamados ch
            LEFT JOIN usuarios u ON ch.criado_por = u.id
            LEFT JOIN usuarios t ON ch.tecnico_id = t.id
            LEFT JOIN filiais f ON ch.filial_id = f.id
            ORDER BY ch.id DESC
        """).fetchall()

    chamados = []
    for r in rows:
        item = dict(r)
        dt = item.get("data_abertura")

        if dt:
            try:
                dt_obj = datetime.fromisoformat(dt)
                item["data_abertura_fmt"] = dt_obj.strftime("%d/%m/%Y %H:%M")
            except:
                item["data_abertura_fmt"] = dt
        else:
            item["data_abertura_fmt"] = "-"

        chamados.append(item)

    conn.close()
    return render_template("chamados.html", chamados=chamados, filiais=get_filiais_for_templates())


@app.route("/chamados/novo", methods=["GET","POST"])
@login_required
def abrir_chamado():
    conn = get_db_connection()

    filiais = conn.execute("SELECT * FROM filiais ORDER BY nome").fetchall()
    tecnicos = conn.execute("SELECT * FROM usuarios WHERE tipo='comprador' ORDER BY nome").fetchall()

    if request.method == "POST":
        data_abertura_raw = request.form.get("data_abertura")
        titulo = request.form.get("titulo","").strip()
        descricao = request.form.get("descricao","").strip()
        prioridade = request.form.get("prioridade") or "Normal"
        filial_id = request.form.get("filial_id") or None

        tecnico_id = request.form.get("tecnico_id")
        if not tecnico_id:
            flash("Selecione um comprador.", "danger")
            return redirect(url_for("abrir_chamado"))

        try:
            if data_abertura_raw:
                date_part = datetime.strptime(data_abertura_raw, "%Y-%m-%d")
                now = datetime.now(pytz.UTC)
                dt = date_part.replace(
                    hour=now.hour, minute=now.minute,
                    second=now.second, microsecond=now.microsecond
                )
            else:
                dt = datetime.now(pytz.UTC)

            data_to_store = dt.isoformat()

        except:
            data_to_store = datetime.now(pytz.UTC).isoformat()

        cur = conn.cursor()
        cur.execute("""INSERT INTO chamados 
            (titulo, descricao, prioridade, criado_por, tecnico_id, filial_id, data_abertura)
            VALUES (?,?,?,?,?,?,?)""",
            (titulo, descricao, prioridade, session["user_id"], tecnico_id, filial_id, data_to_store))
        
        conn.commit()
        chamado_id = cur.lastrowid

        file = request.files.get("anexo")
        if file and file.filename and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            saved_name = safe_unique_filename(filename, "chamado")
            filepath = os.path.join(app.config["UPLOAD_FOLDER"], saved_name)
            file.save(filepath)

            conn.execute("""INSERT INTO anexos 
                (chamado_id, filename, filepath, uploaded_by, uploaded_at) 
                VALUES (?,?,?,?,?)""",
                (chamado_id, filename, saved_name, session["user_id"], data_to_store))
            conn.commit()

        try:
            notif_text = f"Novo requerimento criado por {session.get('nome')} (#{chamado_id})"
            conn.execute("""INSERT INTO notificacoes (mensagem_id, chamado_id, texto, autor_id, autor_nome, lida, criado_em)
                            VALUES (?,?,?,?,?,?,?)""",
                         (None, chamado_id, notif_text, session.get('user_id'),
                          session.get('nome'), 0, data_to_store))
            conn.commit()
        except Exception as e:
            print("Erro ao criar notificação de novo chamado:", e)

        conn.close()
        flash("Chamado criado com sucesso!", "success")
        return redirect(url_for("listar_chamados"))

    conn.close()
    return render_template("abrir_chamado.html", filiais=get_filiais_for_templates(), tecnicos=get_tecnicos_for_templates())

@app.route("/chamados/<int:id>")
@login_required
def visualizar_chamado(id):
    conn = get_db_connection()
    row = conn.execute("""
        SELECT ch.*, u.nome AS criador_nome, t.nome AS tecnico_nome, f.nome AS filial_nome
        FROM chamados ch
        LEFT JOIN usuarios u ON ch.criado_por = u.id
        LEFT JOIN usuarios t ON ch.tecnico_id = t.id
        LEFT JOIN filiais f ON ch.filial_id = f.id
        WHERE ch.id = ?
    """, (id,)).fetchone()

    if not row:
        conn.close()
        flash("Chamado não encontrado.", "danger")
        return redirect(url_for("listar_chamados"))

    ch = dict(row)

    # ---- CONTROLE DE PERMISSÃO ----
    tipo = session.get("tipo")
    user_id = session.get("user_id")

    if tipo == "usuario" and ch.get("criado_por") != user_id:
        conn.close()
        flash("Você não tem permissão para acessar este chamado.", "danger")
        return redirect(url_for("listar_chamados"))

    # ---- FORMATANDO DATAS (COM HORÁRIO BRASIL) ----
    from datetime import datetime, timedelta

    def to_brazil_time(dt):
        return dt - timedelta(hours=3)

    # Data de abertura
    if ch.get("data_abertura"):
        try:
            dt = datetime.fromisoformat(ch["data_abertura"])
            br = to_brazil_time(dt)
            ch["data_abertura_fmt"] = br.strftime("%d/%m/%Y %H:%M")
            ch["data_hora_brasil"] = br.strftime("%d/%m/%Y %H:%M")
        except:
            ch["data_abertura_fmt"] = ch["data_abertura"]
            ch["data_hora_brasil"] = ch["data_abertura"]
    else:
        ch["data_abertura_fmt"] = "-"
        ch["data_hora_brasil"] = "-"

    # Data de resolução
    if ch.get("data_resolucao"):
        try:
            dt2 = datetime.fromisoformat(ch["data_resolucao"])
            br2 = to_brazil_time(dt2)
            ch["data_resolucao_fmt"] = br2.strftime("%d/%m/%Y %H:%M")
        except:
            ch["data_resolucao_fmt"] = ch["data_resolucao"]
    else:
        ch["data_resolucao_fmt"] = "-"

    # ---- ANEXOS E MENSAGENS ----
    anexos = conn.execute(
        "SELECT * FROM anexos WHERE chamado_id=?",
        (id,)
    ).fetchall()

    msgs_raw = conn.execute(
        "SELECT * FROM mensagens WHERE chamado_id=? ORDER BY data ASC",
        (id,)
    ).fetchall()

    mensagens = []
    for m in msgs_raw:
        item = dict(m)
        try:
            dt = datetime.fromisoformat(item["data"])
            br = to_brazil_time(dt)
            item["data_fmt"] = br.strftime("%d/%m/%Y %H:%M")
        except:
            item["data_fmt"] = item["data"]
        mensagens.append(item)

    conn.close()

    return render_template(
        "chamado_detalhe.html",
        chamado=ch,
        anexos=anexos,
        mensagens=mensagens
    )

@app.route("/chamados/<int:id>/mensagem", methods=["POST"])
@login_required
def enviar_mensagem(id):
    texto = request.form.get("texto","").strip()
    file = request.files.get("anexo_msg")

    if not texto and (not file or not file.filename):
        flash("Mensagem vazia.", "warning")
        return redirect(url_for("visualizar_chamado", id=id))

    conn = get_db_connection()
    now = datetime.now(pytz.UTC).isoformat()

    anexo_filename = None
    anexo_path = None

    if file and file.filename and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        saved_name = safe_unique_filename(filename, "msg")
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], saved_name))
        anexo_filename = filename
        anexo_path = saved_name

    cur = conn.cursor()
    cur.execute("""INSERT INTO mensagens 
        (chamado_id, autor_id, autor_nome, autor_tipo, texto, anexo_filename, anexo_path, data) 
        VALUES (?,?,?,?,?,?,?,?)""",
        (id, session["user_id"], session.get("nome"), session.get("tipo"),
         texto, anexo_filename, anexo_path, now))
    conn.commit()
    mensagem_id = cur.lastrowid

    try:
        notif_text = f"{session.get('nome')} comentou no requerimento #{id}: {(texto or '')[:120]}"
        conn.execute("""INSERT INTO notificacoes (mensagem_id, chamado_id, texto, autor_id, autor_nome, lida, criado_em)
                        VALUES (?,?,?,?,?,?,?)""",
                     (mensagem_id, id, notif_text, session.get('user_id'), session.get('nome'), 0, now))
        conn.commit()
    except Exception as e:
        print("Erro ao criar notificação:", e)

    novo_status = request.form.get("novo_status")
    if novo_status and session.get("tipo") in ["tecnico","admin"]:
        conn.execute("UPDATE chamados SET status=?, data_resolucao=? WHERE id=?",
                     (novo_status, datetime.now(pytz.UTC).isoformat() if novo_status == "Concluído" else None, id))
        conn.commit()
        try:
            notif_text = f"Status alterado para {novo_status} por {session.get('nome')} (#{id})"
            conn.execute("""INSERT INTO notificacoes (mensagem_id, chamado_id, texto, autor_id, autor_nome, lida, criado_em)
                            VALUES (?,?,?,?,?,?,?)""",
                         (None, id, notif_text, session.get('user_id'), session.get('nome'), 0, datetime.now(pytz.UTC).isoformat()))
            conn.commit()
        except Exception as e:
            print("Erro notificar status:", e)

    conn.close()
    flash("Mensagem enviada.", "success")

    return redirect(url_for("visualizar_chamado", id=id))


@app.route("/mensagens/<int:chamado_id>/fetch")
@login_required
def fetch_mensagens(chamado_id):
    conn = get_db_connection()
    rows = conn.execute("""
        SELECT id, autor_id, autor_nome, autor_tipo, texto, anexo_filename, anexo_path, data 
        FROM mensagens 
        WHERE chamado_id=? 
        ORDER BY data ASC
    """, (chamado_id,)).fetchall()

    conn.close()
    return jsonify([dict(r) for r in rows])

@app.route("/chamados/<int:id>/status", methods=["POST"])
@login_required
def alterar_status_chamado(id):
    tipo = session.get("tipo")

    # USUÁRIO COMUM INCLUÍDO AQUI
    if tipo not in ["usuario", "tecnico", "admin", "comprador"]:
        return jsonify({"error": "Sem permissão"}), 403

    status = request.form.get("status", "").strip()

    if status not in ["Aberto", "Em andamento", "Concluído"]:
        return jsonify({"error": "Status inválido"}), 400

    conn = get_db_connection()
    conn.execute(
        "UPDATE chamados SET status = ? WHERE id = ?",
        (status, id)
    )
    conn.commit()
    conn.close()

    return jsonify({"status": status}), 200

    flash(f"Status alterado para {novo_status}.", "success")
    return redirect(url_for("visualizar_chamado", id=id))

# ---------- Usuários / Filiais (mantidos) ----------
@app.route("/usuarios")
@login_required
def listar_usuarios():
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    conn = get_db_connection()
    users = conn.execute("""
        SELECT u.*, f.nome as filial_nome 
        FROM usuarios u 
        LEFT JOIN filiais f ON u.filial_id=f.id 
        ORDER BY u.id DESC
    """).fetchall()
    conn.close()

    return render_template("usuarios.html", usuarios=users)


@app.route("/usuarios/novo", methods=["GET","POST"])
@login_required
def novo_usuario():
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    conn = get_db_connection()
    filiais = conn.execute("SELECT * FROM filiais ORDER BY nome").fetchall()

    if request.method == "POST":
        nome = request.form.get("nome")
        usuario = request.form.get("usuario") or None
        email = request.form.get("email")
        senha_raw = request.form.get("senha") or "123"
        senha = generate_password_hash(senha_raw)
        tipo = request.form.get("tipo") or "usuario"
        filial_id = request.form.get("filial_id") or None

        try:
            conn.execute("""INSERT INTO usuarios 
                (nome,usuario,email,senha,tipo,filial_id) 
                VALUES (?,?,?,?,?,?)""",
                (nome, usuario, email, senha, tipo, filial_id))
            conn.commit()

            flash("Usuário criado.", "success")

        except sqlite3.IntegrityError:
            flash("Usuário ou e-mail já existe.", "warning")

        finally:
            conn.close()

        return redirect(url_for("listar_usuarios"))

    conn.close()
    return render_template("usuarios_novo.html", filiais=filiais)


@app.route("/usuarios/editar/<int:id>", methods=["GET","POST"])
@login_required
def editar_usuario(id):
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    conn = get_db_connection()
    usuario = conn.execute("SELECT * FROM usuarios WHERE id=?", (id,)).fetchone()
    filiais = conn.execute("SELECT * FROM filiais ORDER BY nome").fetchall()

    if not usuario:
        conn.close()
        flash("Usuário não encontrado.", "danger")
        return redirect(url_for("listar_usuarios"))

    if request.method == "POST":
        nome = request.form.get("nome")
        usuario_nome = request.form.get("usuario") or None
        email = request.form.get("email")
        tipo = request.form.get("tipo") or "usuario"
        filial_id = request.form.get("filial_id") or None
        senha = request.form.get("senha")

        try:
            if senha:
                senha_hash = generate_password_hash(senha)
                conn.execute("""UPDATE usuarios 
                    SET nome=?, usuario=?, email=?, tipo=?, filial_id=?, senha=? 
                    WHERE id=?""",
                    (nome, usuario_nome, email, tipo, filial_id, senha_hash, id))
            else:
                conn.execute("""UPDATE usuarios 
                    SET nome=?, usuario=?, email=?, tipo=?, filial_id=? 
                    WHERE id=?""",
                    (nome, usuario_nome, email, tipo, filial_id, id))

            conn.commit()
            flash("Usuário atualizado.", "success")

        except sqlite3.IntegrityError:
            flash("Erro: usuário/email duplicado.", "warning")

        finally:
            conn.close()

        return redirect(url_for("listar_usuarios"))

    conn.close()
    return render_template("usuarios_editar.html", usuario=usuario, filiais=filiais)


@app.route("/usuarios/excluir/<int:id>", methods=["POST"])
@login_required
def excluir_usuario(id):
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    conn = get_db_connection()

    if id == session.get("user_id"):
        flash("Você não pode excluir seu próprio usuário.", "warning")
        conn.close()
        return redirect(url_for("listar_usuarios"))

    conn.execute("DELETE FROM usuarios WHERE id=?", (id,))
    conn.commit()
    conn.close()

    flash("Usuário excluído.", "success")
    return redirect(url_for("listar_usuarios"))


@app.route("/filiais")
@login_required
def listar_filiais():
    conn = get_db_connection()
    filiais = conn.execute("SELECT * FROM filiais ORDER BY nome").fetchall()
    conn.close()

    return render_template("filiais.html", filiais=filiais)


@app.route("/filiais/novo", methods=["GET","POST"])
@login_required
def nova_filial():
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    if request.method == "POST":
        nome = request.form.get("nome")
        endereco = request.form.get("endereco")
        cidade = request.form.get("cidade")
        estado = request.form.get("estado")
        telefone = request.form.get("telefone")

        conn = get_db_connection()

        try:
            conn.execute("""INSERT INTO filiais 
                (nome,endereco,cidade,estado,telefone) 
                VALUES (?,?,?,?,?)""",
                (nome, endereco, cidade, estado, telefone))
            conn.commit()

            flash("Filial criada.", "success")

        except sqlite3.IntegrityError:
            flash("Já existe filial com esse nome.", "warning")

        finally:
            conn.close()

        return redirect(url_for("listar_filiais"))

    return render_template("filial_nova.html")


@app.route("/filiais/editar/<int:id>", methods=["GET","POST"])
@login_required
def editar_filial(id):
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_chamados"))

    conn = get_db_connection()
    filial = conn.execute("SELECT * FROM filiais WHERE id=?", (id,)).fetchone()

    if not filial:
        conn.close()
        flash("Filial não encontrada.", "danger")
        return redirect(url_for("listar_filiais"))

    if request.method == "POST":
        nome = request.form.get("nome")
        endereco = request.form.get("endereco")
        cidade = request.form.get("cidade")
        estado = request.form.get("estado")
        telefone = request.form.get("telefone")

        try:
            conn.execute("""UPDATE filiais 
                SET nome=?, endereco=?, cidade=?, estado=?, telefone=? 
                WHERE id=?""",
                (nome, endereco, cidade, estado, telefone, id))
            conn.commit()

            flash("Filial atualizada.", "success")

        except sqlite3.IntegrityError:
            flash("Já existe outra filial com esse nome.", "warning")

        finally:
            conn.close()

        return redirect(url_for("listar_filiais"))

    conn.close()
    return render_template("filial_editar.html", filial=filial)


@app.route("/filiais/excluir/<int:id>", methods=["POST"])
@login_required
def excluir_filial(id):
    if session.get("tipo") != "admin":
        flash("Acesso negado.", "danger")
        return redirect(url_for("listar_filiais"))

    conn = get_db_connection()

    u = conn.execute("SELECT COUNT(*) as cnt FROM usuarios WHERE filial_id=?", (id,)).fetchone()["cnt"]
    c = conn.execute("SELECT COUNT(*) as cnt FROM chamados WHERE filial_id=?", (id,)).fetchone()["cnt"]

    if u > 0 or c > 0:
        flash("Não é possível excluir: existem usuários ou chamados vinculados a esta filial.", "warning")
        conn.close()
        return redirect(url_for("listar_filiais"))

    conn.execute("DELETE FROM filiais WHERE id=?", (id,))
    conn.commit()
    conn.close()

    flash("Filial excluída.", "success")
    return redirect(url_for("listar_filiais"))

# Helper functions to provide template data
def get_filiais_for_templates():
    conn = get_db_connection()
    rows = conn.execute("SELECT id, nome FROM filiais ORDER BY nome").fetchall()
    conn.close()
    return rows

def get_tecnicos_for_templates():
    conn = get_db_connection()
    rows = conn.execute("SELECT * FROM usuarios WHERE tipo='comprador' ORDER BY nome").fetchall()
    conn.close()
    return rows

# Error handler
@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404


# --- TRY IMPORT LOCAL XLSX ON STARTUP (non-blocking) ---
try:
    if LOCAL_XLSX_PATH and os.path.exists(LOCAL_XLSX_PATH):
        print(f"[startup] arquivo local encontrado: {LOCAL_XLSX_PATH} -> iniciando import")
        res = import_products_from_xlsx(LOCAL_XLSX_PATH)
        print("[startup] resultado import:", res)
    elif LOCAL_XLSX_PATH:
        print(f"[startup] arquivo local NÃO encontrado (procura: {LOCAL_XLSX_PATH})")
except Exception as e:
    print("Erro ao tentar importar arquivo local no startup:", e)

@app.route("/acompanhamento_compras", methods=["GET"], endpoint="acompanhamento_compras")
@login_required
def acompanhamento_compras():
    import re
    conn = get_db_connection()

    comprador_id = request.args.get("comprador_id", type=int)

    # =====================================================
    # CARREGA TODOS COMPRADORES PARA O FILTRO
    # =====================================================
    compradores = conn.execute("""
        SELECT id, nome
        FROM usuarios
        WHERE tipo='comprador'
        ORDER BY nome
    """).fetchall()

    # =====================================================
    # 1) CARREGA ITENS SALVOS + STATUS COMPRA (JOIN)
    # =====================================================
    salvos = conn.execute("""
        SELECT 
            ac.id,
            ac.chamado_id,
            ac.codpro,
            ac.titulo,
            ac.descricao,
            ac.quantidade,
            ac.valor_unitario,
            ac.valor_total,
            ac.data_compra,
            ac.nf_compra,
            ac.data_criacao,
            ch.status_compra,
            ch.tecnico_id
        FROM acompanhamento_compras ac
        LEFT JOIN chamados ch ON ch.id = ac.chamado_id
        ORDER BY ac.id ASC
    """).fetchall()

    lista = [dict(r) for r in salvos]

    # Filtra itens pelo comprador se selecionado
    if comprador_id:
        lista = [item for item in lista if item.get("tecnico_id") == comprador_id]

    # Guarda chaves já existentes (para evitar duplicar itens)
    chaves_existentes = set()
    for item in lista:
        codpro = (item.get("codpro") or "").strip().upper()
        descricao = (item.get("descricao") or "").strip().upper()
        chave = f"{item['chamado_id']}_{codpro or descricao}"
        chaves_existentes.add(chave)

    # =====================================================
    # 2) CARREGA CHAMADOS PARA ADICIONAR FALTANTES
    # =====================================================
    chamados_rows = conn.execute("""
        SELECT 
            ch.id,
            ch.titulo,
            ch.descricao,
            ch.status_compra,
            ch.tecnico_id
        FROM chamados ch
        ORDER BY ch.id DESC
    """).fetchall()
    conn.close()

    chamados = [dict(ch) for ch in chamados_rows]

    # =====================================================
    # FUNÇÃO PARA PEGAR QUANTIDADE ORIGINAL DO TÍTULO
    # =====================================================
    def extrair_quantidade(titulo):
        if not titulo:
            return "1 UN"
        partes = titulo.split("-")
        if len(partes) < 2:
            return "1 UN"
        return partes[-1].strip()

    # =====================================================
    # 3) CRIA NOVOS ITENS QUE AINDA NÃO EXISTEM NA TABELA
    # =====================================================
    for ch in chamados:

        # Se filtrando por comprador, ignora chamados de outros compradores
        if comprador_id and ch.get("tecnico_id") != comprador_id:
            continue

        descs = ch["descricao"].split("\n") if ch["descricao"] else []

        for d in descs:
            d = d.strip()
            if not d:
                continue

            # Extrai quantidade do final do texto
            match = re.search(r'(\d+(?:[.,]\d+)?)\s*(PÇ|PC|UN|UNID|M²|M2|M)$', d, re.IGNORECASE)
            if match:
                qtd_completa = f"{match.group(1).replace('.', ',')} {match.group(2).upper()}"
                descricao_sem_qtd = d[:match.start()].strip()
            else:
                qtd_completa = "1 UN"
                descricao_sem_qtd = d

            partes = descricao_sem_qtd.split()
            codpro = partes[0] if partes else ""
            descricao_limpa = " ".join(partes[1:]) if len(partes) > 1 else ""

            chave = f"{ch['id']}_{codpro or descricao_limpa}".strip().upper()

            if chave in chaves_existentes:
                continue
            chaves_existentes.add(chave)

            quantidade_final = extrair_quantidade(ch["titulo"])

            # Adiciona novo item ao acompanhamento (AINDA NÃO SALVO)
            lista.append({
                "id": None,
                "chamado_id": ch["id"],
                "titulo": ch["titulo"],
                "codpro": codpro,
                "descricao": descricao_limpa,
                "quantidade": quantidade_final,
                "valor_unitario": 0.0,
                "valor_total": 0.0,
                "data_compra": "",
                "nf_compra": "",
                "data_criacao": "",
                "status_compra": ch.get("status_compra", ""),
                "tecnico_id": ch.get("tecnico_id")
            })

    # =====================================================
    # 4) RETORNA PARA O HTML COM COMPRADOR
    # =====================================================
    return render_template(
        "acompanhamento_compras.html",
        itens=lista,
        compradores=compradores,
        comprador_id=comprador_id
    )

import sqlite3
from flask import request, jsonify
import time

@app.route("/api/acompanhamento_compras/salvar", methods=["POST"])
@login_required
def api_salvar_acompanhamento():
    if session.get("tipo") not in ("admin", "comprador"):
        return jsonify({"msg": "Acesso negado"}), 403
    dados = request.json
    if not dados:
        return jsonify({"msg": "Nenhum dado recebido"}), 400

    # Converte valor para float de forma segura
    def to_float(val):
        try:
            return float(str(val).replace(",", "."))
        except:
            return 0.0

    # Retry automático para SQLite
    def execute_with_retry(cur, sql, params, retries=5, delay=0.2):
        for i in range(retries):
            try:
                cur.execute(sql, params)
                return
            except sqlite3.OperationalError as e:
                if "database is locked" in str(e):
                    time.sleep(delay)
                else:
                    raise
        raise sqlite3.OperationalError("Banco de dados travado após várias tentativas")

    # Conexão única com timeout
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    try:
        # Inicia transação
        for item in dados:
            item_id = item.get("id")
            chamado_id = item.get("chamado_id")
            codpro = (item.get("codpro") or "").strip().upper()
            quantidade = item.get("quantidade", 0)
            valor_unitario = to_float(item.get("valor_unitario"))
            valor_total = to_float(item.get("valor_total"))
            titulo = item.get("titulo") or ""
            descricao = item.get("descricao", None)

            data_compra = item.get("data_compra") or None
            if data_compra and data_compra.strip() == "":
                data_compra = None

            nf_compra = item.get("nf_compra") or None
            if nf_compra and nf_compra.strip() == "":
                nf_compra = None

            status_compra = (item.get("status_compra") or "").strip()

            if item_id:
                # UPDATE
                execute_with_retry(cur, """
                    UPDATE acompanhamento_compras
                    SET 
                        titulo=?,
                        descricao=COALESCE(?, descricao),
                        quantidade=?,
                        valor_unitario=?,
                        valor_total=?,
                        data_compra=?,
                        nf_compra=?,
                        atualizado_em=DATE('now')
                    WHERE id=?
                """, (
                    titulo, descricao, quantidade, valor_unitario, valor_total,
                    data_compra, nf_compra, item_id
                ))
            else:
                # INSERT
                execute_with_retry(cur, """
                    INSERT INTO acompanhamento_compras
                    (chamado_id, titulo, codpro, descricao, quantidade, valor_unitario,
                     valor_total, data_compra, nf_compra, data_criacao)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, DATE('now'))
                """, (
                    chamado_id, titulo, codpro, descricao, quantidade,
                    valor_unitario, valor_total, data_compra, nf_compra
                ))
                item_id = cur.lastrowid

            # Atualiza status do chamado, se necessário
            if chamado_id and status_compra != "":
                execute_with_retry(cur, """
                    UPDATE chamados
                    SET status_compra=?
                    WHERE id=?
                """, (status_compra, chamado_id))

        # Commit único para toda a requisição
        conn.commit()
        return jsonify({"msg": "Dados salvos com sucesso!"}), 200

    except sqlite3.OperationalError as e:
        return jsonify({"msg": "Erro ao salvar dados", "erro": str(e)}), 500

    finally:
        cur.close()
        conn.close()

# ---------------------------
# ROTA PARA ALTERAR STATUS DE COMPRA (VERSÃO ROBUSTA)
# ---------------------------
@app.route("/chamados/<int:id>/status_compra", methods=["POST"])
@login_required
def alterar_status_compra(id):
    # só comprador pode alterar
    if session.get("tipo") != "comprador":
        return jsonify({"ok": False, "erro": "Acesso negado"}), 403

    # aceita tanto form-urlencoded quanto json
    if request.is_json:
        recebido = (request.get_json(silent=True) or {}).get("status_compra", "")
    else:
        recebido = request.form.get("status_compra", "")

    recebido = (recebido or "").strip()

    # valida valores permitidos (mesmos do front)
    permitidos = ["", "Vazio", "Troca", "Bonificado", "Não Bonificado"]
    if recebido not in permitidos:
        return jsonify({"ok": False, "erro": f"Status inválido: '{recebido}'", "permitidos": permitidos}), 400

    try:
        conn = get_db_connection()
        cur = conn.cursor()

        # verifica se o chamado existe
        existe = cur.execute("SELECT id FROM chamados WHERE id=?", (id,)).fetchone()
        if not existe:
            conn.close()
            return jsonify({"ok": False, "erro": "Chamado não encontrado"}), 404

        # atualiza
        cur.execute("UPDATE chamados SET status_compra=?, data_resolucao = data_resolucao WHERE id=?",
                    (recebido, id))
        conn.commit()
        conn.close()

        # opcional: cria notificação
        try:
            criar_notificacao(id, f"Status de compra alterado para '{recebido}' por {session.get('nome')}", session.get('user_id'), session.get('nome'))
        except Exception:
            pass

        return jsonify({"ok": True, "novo_status": recebido})

    except Exception as e:
        # log no terminal para debug
        print("[erro alterar_status_compra]", str(e))
        try:
            conn.close()
        except:
            pass
        return jsonify({"ok": False, "erro": "Erro interno", "detalhe": str(e)}), 500

from functools import wraps
from werkzeug.utils import secure_filename
from datetime import datetime
from decimal import Decimal

import sqlite3
import os
import json
import uuid

from flask import (
    render_template,
    request,
    session,
    flash,
    redirect,
    url_for,
    jsonify
)

# =========================================================
# CONFIG
# =========================================================

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

STATUS_ABERTO = "ABERTO"
STATUS_ANALISE = "EM ANALISE"
STATUS_FINALIZADO = "FINALIZADO"

os.makedirs(
    UPLOAD_FOLDER,
    exist_ok=True
)

# =========================================================
# DB
# =========================================================

def get_db_connection():

    conn = sqlite3.connect(DB_PATH)

    conn.row_factory = sqlite3.Row

    return conn

# =========================================================
# LOGIN REQUIRED
# =========================================================

def login_required(f):

    @wraps(f)

    def decorated(*args, **kwargs):

        if not session.get("user_id") and not session.get("usuario_id") and not session.get("id"):

            flash(
                "Faça login para continuar.",
                "danger"
            )

            return redirect(
                url_for("login")
            )

        return f(*args, **kwargs)

    return decorated

# =========================================================
# USUARIO LOGADO
# =========================================================

def obter_usuario_id_logado():

    usuario_id = (
        session.get("usuario_id")
        or session.get("user_id")
        or session.get("id")
    )

    return usuario_id

# =========================================================
# PADRONIZAR FILIAL
# =========================================================

def padronizar_filial(filial):

    filial = (
        str(filial or "")
        .strip()
        .upper()
        .replace("  ", " ")
    )

    mapa_filiais = {

        "F51": "F51-MATRIZ",
        "F51 MATRIZ": "F51-MATRIZ",
        "F51- MATRIZ": "F51-MATRIZ",
        "F51-MATRIZ": "F51-MATRIZ",

        "F52": "F52-CD",
        "F52 CD": "F52-CD",
        "F52- CD": "F52-CD",
        "F52-CD": "F52-CD",
        "CD": "F52-CD",

        "F53": "F53-OURO BRANCO",
        "F53 OURO BRANCO": "F53-OURO BRANCO",
        "F53- OURO BRANCO": "F53-OURO BRANCO",
        "F53-OURO BRANCO": "F53-OURO BRANCO",

        "F54": "F54-CONGONHAS",
        "F54 CONGONHAS": "F54-CONGONHAS",
        "F54- CONGONHAS": "F54-CONGONHAS",
        "F54-CONGONHAS": "F54-CONGONHAS"

    }

    return mapa_filiais.get(
        filial,
        filial
    )

# =========================================================
# ALIASES FILIAL
# =========================================================

def aliases_filial(filial):

    filial = padronizar_filial(filial)

    mapa_aliases = {

        "F51-MATRIZ": [
            "F51-MATRIZ",
            "F51- MATRIZ",
            "F51 MATRIZ",
            "F51"
        ],

        "F52-CD": [
            "F52-CD",
            "F52- CD",
            "F52 CD",
            "F52",
            "CD"
        ],

        "F53-OURO BRANCO": [
            "F53-OURO BRANCO",
            "F53- OURO BRANCO",
            "F53 OURO BRANCO",
            "F53"
        ],

        "F54-CONGONHAS": [
            "F54-CONGONHAS",
            "F54- CONGONHAS",
            "F54 CONGONHAS",
            "F54"
        ]

    }

    return mapa_aliases.get(
        filial,
        [filial]
    )

# =========================================================
# BUSCAR FILIAL DO USUARIO LOGADO
# =========================================================

def obter_filial_usuario_logado(conn):

    usuario_id = obter_usuario_id_logado()

    if not usuario_id:

        return ""

    usuario = conn.execute("""

        SELECT
            u.id,
            u.nome,
            u.filial_id,
            f.nome AS filial_nome

        FROM usuarios u

        LEFT JOIN filiais f
            ON f.id = u.filial_id

        WHERE u.id = ?

    """, (

        usuario_id,

    )).fetchone()

    if not usuario:

        return ""

    filial_usuario = (
        usuario["filial_nome"]
        or ""
    )

    return padronizar_filial(
        filial_usuario
    )

# =========================================================
# NORMALIZAR FURO
# =========================================================

def normalizar_furo(row):

    item = dict(row)

    item["descricao"] = (
        item.get("descricao")
        or "-"
    )

    item["local_fisico"] = (
        item.get("local_fisico")
        or "-"
    )

    item["conferido_por_nome"] = (
        item.get("conferido_por_nome")
        or "-"
    )

    item["ajustado_por_nome"] = (
        item.get("ajustado_por_nome")
        or item.get("usuario_nome")
        or "-"
    )

    item["filial"] = padronizar_filial(
        item.get("filial")
    )

    item["status"] = (
        item.get("status")
        or STATUS_ABERTO
    )

    try:

        qtde_encontrada = Decimal(
            str(item.get("qtde_encontrada") or 0)
        )

    except:

        qtde_encontrada = Decimal("0")

    try:

        qtde_sistema = Decimal(
            str(item.get("qtde_sistema") or 0)
        )

    except:

        qtde_sistema = Decimal("0")

    item["qtde_encontrada"] = float(
        qtde_encontrada
    )

    item["qtde_sistema"] = float(
        qtde_sistema
    )

    diferenca = (
        qtde_encontrada
        -
        qtde_sistema
    )

    item["dif"] = float(
        diferenca
    )

    item["diferenca"] = float(
        diferenca
    )

    item["status_visual"] = (
        "CERTO"
        if diferenca == 0
        else "ERRADO"
    )

    item["status_erp"] = (
        item.get("status_erp")
        or "PENDENTE"
    )

    data_furo = item.get("data_furo")

    if data_furo:

        try:

            dt = datetime.fromisoformat(
                str(data_furo)
            )

            item["data_furo_fmt"] = dt.strftime(
                "%d/%m/%Y"
            )

        except:

            item["data_furo_fmt"] = str(
                data_furo
            )

    else:

        item["data_furo_fmt"] = "-"

    return item

# =========================================================
# DASHBOARD AJUSTE ESTOQUE ERP
# =========================================================

@app.route("/furo_estoque", methods=["GET"])
@login_required
def furo_estoque():

    conn = get_db_connection()

    # ================= USUARIO / FILIAL LOGADA =================

    filial_usuario = obter_filial_usuario_logado(conn)

    # ================= FILTROS =================

    busca = request.args.get("busca", "").strip()

    filial = request.args.get("filial", "").strip()

    patio = request.args.get("patio", "").strip()

    status = request.args.get("status", "").strip()

    conferido_por = request.args.get("conferido_por", "").strip()

    tipo_dif = request.args.get("tipo_dif", "").strip()

    data_inicio = request.args.get("data_inicio", "").strip()

    data_fim = request.args.get("data_fim", "").strip()

    if not filial:

        filial = filial_usuario

    filial = padronizar_filial(
        filial
    )

    # ================= QUERY BASE =================

    query = """

        SELECT
            f.*,
            u.nome AS conferido_por_nome

        FROM furo_estoque f

        LEFT JOIN usuarios u
            ON u.id = f.conferido_por

        WHERE 1=1

    """

    params = []

    # ================= FILTROS SQL =================

    if busca:

        query += """

            AND (
                f.descricao LIKE ?
                OR f.codpro LIKE ?
            )

        """

        params.extend([
            f"%{busca}%",
            f"%{busca}%"
        ])

    if filial:

        filiais_possiveis = aliases_filial(
            filial
        )

        placeholders = ",".join(
            ["?"] * len(filiais_possiveis)
        )

        query += f"""

            AND f.filial IN ({placeholders})

        """

        params.extend(
            filiais_possiveis
        )

    if patio:

        query += """

            AND f.patio = ?

        """

        params.append(
            patio
        )

    if status:

        query += """

            AND COALESCE(f.status, ?) = ?

        """

        params.extend([
            STATUS_ABERTO,
            status
        ])

    if conferido_por:

        query += """

            AND f.conferido_por = ?

        """

        params.append(
            conferido_por
        )

    if data_inicio:

        query += """

            AND date(f.data_furo) >= date(?)

        """

        params.append(
            data_inicio
        )

    if data_fim:

        query += """

            AND date(f.data_furo) <= date(?)

        """

        params.append(
            data_fim
        )

    query += """

        ORDER BY f.id DESC

    """

    rows = conn.execute(
        query,
        params
    ).fetchall()

    # ================= DIFERENÇA / KPIS =================

    itens = []

    total_ajustes = 0
    total_corretos = 0
    total_errados = 0

    for row in rows:

        item = normalizar_furo(
            row
        )

        if not item.get("filial"):

            item["filial"] = filial_usuario

        item["filial"] = padronizar_filial(
            item.get("filial")
        )

        item["status"] = (
            item.get("status")
            or STATUS_ABERTO
        )

        dif = item.get("dif") or 0

        if tipo_dif == "ok" and dif != 0:

            continue

        if tipo_dif == "pos" and dif <= 0:

            continue

        if tipo_dif == "neg" and dif >= 0:

            continue

        total_ajustes += 1

        if dif == 0:

            total_corretos += 1

        else:

            total_errados += 1

        itens.append(
            item
        )

    # ================= USUARIOS =================

    usuarios = conn.execute("""

        SELECT
            id,
            nome

        FROM usuarios

        ORDER BY nome

    """).fetchall()

    conn.close()

    return render_template(

        "furo_estoque.html",

        itens=itens,

        usuarios=usuarios,

        busca=busca,

        filial=filial,

        filial_usuario=filial_usuario,

        patio=patio,

        status=status,

        conferido_por=conferido_por,

        tipo_dif=tipo_dif,

        data_inicio=data_inicio,

        data_fim=data_fim,

        total_ajustes=total_ajustes,

        total_corretos=total_corretos,

        total_errados=total_errados

    )

# =========================================================
# NOVO AJUSTE ESTOQUE ERP
# =========================================================

# =========================================================
# NOVO AJUSTE ESTOQUE ERP
# =========================================================

@app.route("/novo_furo", methods=["GET", "POST"])
@login_required
def novo_furo():

    conn = get_db_connection()

    cursor = conn.cursor()

    # =========================================
    # GARANTIR COLUNAS NECESSARIAS
    # =========================================

    colunas_furo = cursor.execute("""
        PRAGMA table_info(furo_estoque)
    """).fetchall()

    nomes_colunas = [
        coluna["name"]
        for coluna in colunas_furo
    ]

    if "codpro" not in nomes_colunas:

        cursor.execute("""
            ALTER TABLE furo_estoque
            ADD COLUMN codpro TEXT
        """)

        conn.commit()

    if "status" not in nomes_colunas:

        cursor.execute("""
            ALTER TABLE furo_estoque
            ADD COLUMN status TEXT DEFAULT 'ABERTO'
        """)

        conn.commit()

    # =========================================
    # USUARIO LOGADO
    # =========================================

    usuario_id = (
        session.get("usuario_id")
        or session.get("user_id")
        or session.get("id")
    )

    if not usuario_id:

        conn.close()

        flash(
            "Sessão expirada. Faça login novamente.",
            "error"
        )

        return redirect(
            url_for("login")
        )

    # =========================================
    # BUSCA USUARIO + FILIAL
    # =========================================

    usuario = cursor.execute("""

        SELECT
            u.id,
            u.nome,
            u.filial_id,
            f.nome AS filial_nome

        FROM usuarios u

        LEFT JOIN filiais f
            ON f.id = u.filial_id

        WHERE u.id = ?

    """, (

        usuario_id,

    )).fetchone()

    # =========================================
    # FILIAL
    # =========================================

    filial_usuario = ""

    if usuario:

        filial_usuario = (
            usuario["filial_nome"]
            or ""
        ).strip().upper()

    # =========================================
    # PADRONIZA FILIAIS
    # =========================================

    mapa_filiais = {

        "F51": "F51-MATRIZ",
        "F51 MATRIZ": "F51-MATRIZ",
        "F51- MATRIZ": "F51-MATRIZ",
        "F51-MATRIZ": "F51-MATRIZ",

        "F52": "F52-CD",
        "F52 CD": "F52-CD",
        "F52- CD": "F52-CD",
        "F52-CD": "F52-CD",
        "CD": "F52-CD",

        "F53": "F53-OURO BRANCO",
        "F53 OURO BRANCO": "F53-OURO BRANCO",
        "F53- OURO BRANCO": "F53-OURO BRANCO",
        "F53-OURO BRANCO": "F53-OURO BRANCO",

        "F54": "F54-CONGONHAS",
        "F54 CONGONHAS": "F54-CONGONHAS",
        "F54- CONGONHAS": "F54-CONGONHAS",
        "F54-CONGONHAS": "F54-CONGONHAS"

    }

    filial_usuario = mapa_filiais.get(
        filial_usuario,
        filial_usuario
    )

    # =========================================
    # USUARIOS
    # =========================================

    usuarios = cursor.execute("""

        SELECT
            id,
            nome

        FROM usuarios

        ORDER BY nome

    """).fetchall()

    # =========================================
    # POST
    # =========================================

    if request.method == "POST":

        try:

            filial = filial_usuario

            codpro = request.form.get(
                "codpro",
                ""
            ).strip()

            patio = request.form.get(
                "patio",
                ""
            ).strip()

            local_fisico = request.form.get(
                "local_fisico",
                ""
            ).strip()

            data_furo = datetime.now().strftime(
                "%Y-%m-%d"
            )

            descricao = request.form.get(
                "descricao",
                ""
            ).strip()

            # =========================================
            # SE VEIO "CODIGO - DESCRICAO", SEPARA
            # =========================================

            if descricao and " - " in descricao:

                partes_descricao = descricao.split(
                    " - ",
                    1
                )

                codigo_digitado = partes_descricao[0].strip()

                descricao_digitada = partes_descricao[1].strip()

                if codigo_digitado and not codpro:

                    codpro = codigo_digitado

                if descricao_digitada:

                    descricao = descricao_digitada

            lote = request.form.get(
                "lote",
                ""
            ).strip()

            tonalidade = request.form.get(
                "tonalidade",
                ""
            ).strip()

            bitola = request.form.get(
                "bitola",
                ""
            ).strip()

            qtde_encontrada = request.form.get(
                "qtde_encontrada",
                0
            )

            qtde_sistema = request.form.get(
                "qtde_sistema",
                0
            )

            conferido_por = request.form.get(
                "conferido_por"
            )

            cursor.execute("""

                INSERT INTO furo_estoque (

                    codpro,
                    filial,
                    patio,
                    local_fisico,
                    data_furo,
                    descricao,
                    lote,
                    tonalidade,
                    bitola,
                    qtde_encontrada,
                    qtde_sistema,
                    conferido_por,
                    usuario_id,
                    status

                )

                VALUES (

                    ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?

                )

            """, (

                codpro,
                filial,
                patio,
                local_fisico,
                data_furo,
                descricao,
                lote,
                tonalidade,
                bitola,
                qtde_encontrada,
                qtde_sistema,
                conferido_por,
                usuario_id,
                "ABERTO"

            ))

            conn.commit()

            flash(
                "Ajuste salvo com sucesso!",
                "success"
            )

            conn.close()

            return redirect(
                url_for("novo_furo")
            )

        except Exception as e:

            conn.rollback()

            flash(
                f"Erro ao salvar ajuste: {e}",
                "error"
            )

    conn.close()

    return render_template(

        "novo_furo.html",

        usuarios=usuarios,

        filial_usuario=filial_usuario

    )

@app.route("/furo/<int:furo_id>", methods=["GET", "POST"], endpoint="ver_furo")
@login_required
def ver_furo(furo_id):

    conn = get_db_connection()

    row = conn.execute("""

        SELECT
            f.*,
            uc.nome AS conferido_por_nome,
            ua.nome AS ajustado_por_nome

        FROM furo_estoque f

        LEFT JOIN usuarios uc
            ON uc.id = f.conferido_por

        LEFT JOIN usuarios ua
            ON ua.id = f.ajustado_por

        WHERE f.id = ?

    """, (

        furo_id,

    )).fetchone()

    if not row:

        conn.close()

        flash(
            "Ajuste não encontrado.",
            "danger"
        )

        return redirect(
            url_for("furo_estoque")
        )

    furo = dict(row)

    if not furo.get("status"):

        furo["status"] = "ABERTO"

    try:

        mensagens = json.loads(
            furo.get("mensagens") or "[]"
        )

    except:

        mensagens = []

    if request.method == "POST":

        try:

            nova_msg = request.form.get(
                "mensagens",
                ""
            ).strip()

            novo_status = request.form.get(
                "status",
                "ABERTO"
            ).strip()

            if novo_status not in [
                "ABERTO",
                "EM ANALISE",
                "FINALIZADO"
            ]:

                novo_status = "ABERTO"

            usuario_nome = (
                session.get("nome")
                or session.get("usuario_nome")
                or "Usuário"
            )

            usuario_id = (
                session.get("usuario_id")
                or session.get("user_id")
                or session.get("id")
            )

            if nova_msg:

                mensagens.append({

                    "usuario": usuario_nome,

                    "texto": nova_msg,

                    "datahora": datetime.now().strftime(
                        "%d/%m/%Y %H:%M"
                    )

                })

            ajustado_por = furo.get("ajustado_por")

            data_ajuste_erp = furo.get("data_ajuste_erp")

            if novo_status == "FINALIZADO":

                ajustado_por = usuario_id

                data_ajuste_erp = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S"
                )

            conn.execute("""

                UPDATE furo_estoque

                SET
                    mensagens = ?,
                    status = ?,
                    ajustado_por = ?,
                    data_ajuste_erp = ?

                WHERE id = ?

            """, (

                json.dumps(
                    mensagens,
                    ensure_ascii=False
                ),

                novo_status,

                ajustado_por,

                data_ajuste_erp,

                furo_id

            ))

            conn.commit()

            flash(
                "Ajuste atualizado com sucesso!",
                "success"
            )

            conn.close()

            return redirect(
                url_for(
                    "ver_furo",
                    furo_id=furo_id
                )
            )

        except Exception as e:

            conn.rollback()

            conn.close()

            flash(
                f"Erro ao atualizar ajuste: {e}",
                "danger"
            )

            return redirect(
                url_for(
                    "ver_furo",
                    furo_id=furo_id
                )
            )

    data_erp_fmt = "-"

    if furo.get("data_ajuste_erp"):

        try:

            dt = datetime.fromisoformat(
                str(furo["data_ajuste_erp"])
            )

            data_erp_fmt = dt.strftime(
                "%d/%m/%Y %H:%M"
            )

        except:

            data_erp_fmt = furo["data_ajuste_erp"]

    furo["data_ajuste_erp_fmt"] = data_erp_fmt

    data_furo_fmt = "-"

    if furo.get("data_furo"):

        try:

            dt = datetime.fromisoformat(
                str(furo["data_furo"])
            )

            data_furo_fmt = dt.strftime(
                "%d/%m/%Y"
            )

        except:

            data_furo_fmt = furo["data_furo"]

    furo["data_furo_fmt"] = data_furo_fmt

    furo["mensagens"] = mensagens

    furo["conferido_por_nome"] = (
        furo.get("conferido_por_nome")
        or "-"
    )

    furo["ajustado_por_nome"] = (
        furo.get("ajustado_por_nome")
        or "-"
    )

    conn.close()

    return render_template(

        "ver_furo.html",

        furo=furo

    )

from datetime import datetime, timedelta
from decimal import Decimal
import os
try:
    import pyodbc
except ImportError:
    pyodbc = None
import pytz

from flask import (
    Flask, render_template, request, session,
    flash, redirect, url_for, jsonify, send_from_directory
)

from werkzeug.utils import secure_filename

# =========================================================
# CONSTANTES DE STATUS
# =========================================================

STATUS_PENDENTE = "Pendente"
STATUS_CONFERIDO = "Conferido"
STATUS_OK = "Sem Divergência"
STATUS_DIVERGENTE = "Divergente"

# =========================================================
# INICIALIZAÇÃO TABELAS LOCAIS
# =========================================================

def init_status_conferencia():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS status_conferencia (
            numped INTEGER PRIMARY KEY,
            status TEXT DEFAULT 'Pendente'
        )
    """)
    conn.commit()
    conn.close()

def init_conferencias():
    conn = get_db_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS conferencias (
            numped INTEGER,
            codpro TEXT,
            qtd_contada REAL,
            usuario TEXT,
            tonalidade_bitola TEXT,
            enderecamento TEXT,
            data_chegada TEXT,
            PRIMARY KEY (numped, codpro)
        )
    """)
    conn.commit()
    conn.close()

init_status_conferencia()
init_conferencias()

# =========================================================
# CONEXÃO ERP
# =========================================================

def get_erp_connection():
    if pyodbc is None:
        raise RuntimeError("Dependência pyodbc não instalada. Instale as dependências de produção e o driver ODBC.")

    # Dados fixos do ambiente conforme solicitado (Hardcoded para funcionamento imediato)
    # Recomenda-se o uso de variáveis de ambiente para maior segurança em produção.
    driver = os.environ.get("ERP_ODBC_DRIVER", "ODBC Driver 17 for SQL Server")
    server = os.environ.get("ERP_DB_SERVER", "192.168.3.32")
    database = os.environ.get("ERP_DB_NAME", "BDENTER")
    user = os.environ.get("ERP_DB_USER", "microuni")
    password = os.environ.get("ERP_DB_PASSWORD", "microuni")

    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        f"DATABASE={database};"
        f"UID={user};"
        f"PWD={password};"
        "TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str)

# =========================================================
# ROTA LISTAGEM DE PEDIDOS
# =========================================================

@app.route("/conferencia")
@login_required
def conferencia():

    from datetime import datetime, timedelta

    user_tipo = session.get("tipo")
    user_is_admin = user_tipo in ("admin", "comprador")

    pedido = request.args.get("pedido", "").strip()
    fornecedor = request.args.get("fornecedor", "").strip().lower()
    produto = request.args.get("produto", "").strip()
    status_filtro = request.args.get("status", "").strip()
    sort = request.args.get("sort", "data_desc")

    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    previsao_inicio = request.args.get("previsao_inicio")
    previsao_fim = request.args.get("previsao_fim")

    # 🔥 NOVO: filtro rápido
    filtro_rapido = request.args.get("previsao_rapida")

    hoje = datetime.today().date()

    if not data_inicio or not data_fim:
        data_fim = hoje.strftime('%Y-%m-%d')
        data_inicio = (hoje - timedelta(days=30)).strftime('%Y-%m-%d')

    # 🔥 aplica filtros rápidos
    if filtro_rapido == "hoje":
        previsao_inicio = previsao_fim = hoje.strftime('%Y-%m-%d')

    elif filtro_rapido == "amanha":
        amanha = hoje + timedelta(days=1)
        previsao_inicio = previsao_fim = amanha.strftime('%Y-%m-%d')

    elif filtro_rapido == "semana":
        inicio_semana = hoje
        fim_semana = hoje + timedelta(days=7)
        previsao_inicio = inicio_semana.strftime('%Y-%m-%d')
        previsao_fim = fim_semana.strftime('%Y-%m-%d')

    order_by = "DATA_PEDIDO DESC"
    if sort == "pedido_desc":
        order_by = "NUMPED DESC"
    elif sort == "pedido_asc":
        order_by = "NUMPED ASC"
    elif sort == "data_asc":
        order_by = "DATA_PEDIDO ASC"
    elif sort == "previsao_asc":
        order_by = "P.DTPREVREC ASC"
    elif sort == "previsao_desc":
        order_by = "P.DTPREVREC DESC"

    pedidos_com_status = []

    try:
        with get_erp_connection() as erp_conn:
            cursor = erp_conn.cursor()

            query = """
                SELECT DISTINCT 
                    V.NUMPED, 
                    V.DATA_PEDIDO, 
                    V.FORNECEDOR,
                    P.DTPREVREC
                FROM _VISAO_CARGA_PEDIDO_COMPRA V
                LEFT JOIN PEDIFORCAD P 
                    ON P.NUMPED = V.NUMPED
                WHERE 1=1
            """

            params = []

            query += " AND CAST(V.DATA_PEDIDO AS DATE) BETWEEN ? AND ?"
            params.extend([data_inicio, data_fim])

            # 🔥 PREVISÃO ENTREGA
            if previsao_inicio and previsao_fim:
                query += " AND CAST(P.DTPREVREC AS DATE) BETWEEN ? AND ?"
                params.extend([previsao_inicio, previsao_fim])

            # 🔥 atrasados
            if filtro_rapido == "atrasados":
                query += " AND CAST(P.DTPREVREC AS DATE) < ?"
                params.append(hoje.strftime('%Y-%m-%d'))

            if pedido:
                query += " AND V.NUMPED = ?"
                params.append(pedido)

            if fornecedor:
                query += " AND LOWER(V.FORNECEDOR) LIKE ?"
                params.append(f"%{fornecedor}%")

            if produto:
                query += """
                    AND (
                        CAST(V.CODPRO AS VARCHAR) LIKE ?
                        OR UPPER(V.PRODUTO) LIKE UPPER(?)
                    )
                """
                like = f"%{produto}%"
                params.extend([like, like])

            query += f" ORDER BY {order_by}"

            cursor.execute(query, params)
            pedidos = cursor.fetchall()

        with get_db_connection() as conn:

            status_map = {
                r["numped"]: r["status"]
                for r in conn.execute("SELECT numped, status FROM status_conferencia")
            }

            chegada_map = {
                r["numped"]: r["data_chegada"]
                for r in conn.execute("""
                    SELECT numped, MAX(data_chegada) as data_chegada
                    FROM conferencias
                    GROUP BY numped
                """)
            }

            usuario_map = {
                r["numped"]: r["usuario"]
                for r in conn.execute("""
                    SELECT numped, GROUP_CONCAT(DISTINCT usuario) as usuario
                    FROM conferencias
                    GROUP BY numped
                """)
            }

        for p in pedidos:

            status_atual = status_map.get(p.NUMPED, STATUS_PENDENTE)
            data_chegada = chegada_map.get(p.NUMPED)
            usuario_conf = usuario_map.get(p.NUMPED)

            if status_filtro and status_atual != status_filtro:
                continue

            previsao_data = p.DTPREVREC

            atrasado = False
            if previsao_data:
                try:
                    atrasado = previsao_data.date() < hoje
                except:
                    pass

            pedidos_com_status.append({
                "NUMPED": p.NUMPED,
                "DATA_PEDIDO": p.DATA_PEDIDO.strftime("%d/%m/%Y"),
                "FORNECEDOR": p.FORNECEDOR,
                "PREVISAO_ENTREGA": previsao_data.strftime("%d/%m/%Y") if previsao_data else None,
                "ATRASADO": atrasado,
                "STATUS": status_atual,
                "DATA_CHEGADA": datetime.strptime(
                    data_chegada, "%Y-%m-%d"
                ).strftime("%d/%m/%Y") if data_chegada else None,
                "USUARIO_CONFERENCIA": usuario_conf if usuario_conf else "-"
            })

    except Exception as e:
        flash(f"Erro ao carregar pedidos: {e}", "danger")

    return render_template(
        "conferencia.html",
        pedidos=pedidos_com_status,
        pedido=pedido,
        fornecedor=fornecedor,
        produto=produto,
        data_inicio=data_inicio,
        data_fim=data_fim,
        previsao_inicio=previsao_inicio,
        previsao_fim=previsao_fim,
        sort=sort,
        status_filtro=status_filtro,
        user_is_admin=user_is_admin,
        user_tipo=user_tipo
    )

# =========================================================
# SALVAR STATUS
# =========================================================

@app.route("/conferencia/salvar_status", methods=["POST"])
@login_required
def salvar_status():

    numped = request.form.get("numped", type=int)
    novo_status = request.form.get("status", "").strip()
    user_tipo = session.get("tipo")

    if not numped or not novo_status:
        return jsonify({"sucesso": False, "erro": "Dados inválidos"}), 400

    try:
        with get_db_connection() as conn:
            atual = conn.execute("SELECT status FROM status_conferencia WHERE numped = ?", (numped,)).fetchone()
            status_atual = atual["status"] if atual else STATUS_PENDENTE

            # Regras de permissão
            if user_tipo == "usuario":
                if not (status_atual == STATUS_PENDENTE and novo_status == STATUS_CONFERIDO):
                    return jsonify({"sucesso": False, "erro": "Ação não permitida"}), 403
            elif user_tipo == "comprador":
                if not (status_atual == STATUS_CONFERIDO and novo_status in (STATUS_OK, STATUS_DIVERGENTE)):
                    return jsonify({"sucesso": False, "erro": "Ação não permitida"}), 403
            elif user_tipo == "admin":
                pass
            else:
                return jsonify({"sucesso": False, "erro": "Perfil inválido"}), 403

            conn.execute("""
                INSERT INTO status_conferencia (numped, status)
                VALUES (?, ?)
                ON CONFLICT(numped)
                DO UPDATE SET status = excluded.status
            """, (numped, novo_status))

            conn.commit()

        return jsonify({"sucesso": True})

    except Exception as e:
        return jsonify({"sucesso": False, "erro": str(e)}), 500

# =========================================================
# CONFERÊNCIA DE ITENS DO PEDIDO
# =========================================================

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")

@app.route("/conferencia/<int:num_pedido>", methods=["GET", "POST"])
@login_required
def conferencia_itens(num_pedido):

    user_tipo = session.get("tipo")
    usuario = session.get("nome", "desconhecido")
    autor_id = session.get("user_id", 0)

    if user_tipo not in ("usuario", "admin", "comprador"):
        flash("Acesso não autorizado.", "warning")
        return redirect(url_for("login"))

    # ================= ERP =================
    try:

        with get_erp_connection() as erp_conn:

            cursor = erp_conn.cursor()

            cursor.execute("""
                SELECT 
                    CODPRO,
                    PRODUTO,
                    QTDE,
                    TOTAL_PRECO_FINAL
                FROM _VISAO_CARGA_PEDIDO_COMPRA
                WHERE NUMPED = ?
                ORDER BY CODPRO
            """, (num_pedido,))

            itens = [
                dict(zip([c[0] for c in cursor.description], r))
                for r in cursor.fetchall()
            ]

    except Exception as e:

        flash(f"Erro ERP: {e}", "danger")
        itens = []

    # ================= POST =================
    if request.method == "POST":

        try:

            with get_db_connection() as conn:

                # =====================================================
                # SALVAR ITENS
                # =====================================================

                for item in itens:

                    codpro = item["CODPRO"]

                    qtd = request.form.get(f"qtd_{codpro}")
                    tonalidade = request.form.get(f"tonalidade_{codpro}")
                    enderecamento = request.form.get(f"enderecamento_{codpro}")
                    data_chegada = request.form.get(f"data_chegada_{codpro}")

                    lote = request.form.get(f"lote_{codpro}")
                    peso = request.form.get(f"peso_{codpro}")
                    pei = request.form.get(f"pei_{codpro}")
                    area_m2 = request.form.get(f"area_m2_{codpro}")

                    qtd_final = (
                        float(qtd.replace(",", "."))
                        if qtd else None
                    )

                    if any([
                        qtd_final,
                        tonalidade,
                        enderecamento,
                        data_chegada,
                        lote,
                        peso,
                        pei,
                        area_m2
                    ]):

                        conn.execute("""
                            DELETE FROM conferencias
                            WHERE numped = ?
                            AND codpro = ?
                        """, (
                            num_pedido,
                            codpro
                        ))

                        conn.execute("""
                            INSERT INTO conferencias (
                                numped,
                                codpro,
                                qtd_contada,
                                usuario,
                                tonalidade_bitola,
                                enderecamento,
                                data_chegada,
                                lote,
                                peso,
                                pei,
                                area_m2
                            )
                            VALUES (
                                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?
                            )
                        """, (
                            num_pedido,
                            codpro,
                            qtd_final,
                            usuario,
                            tonalidade,
                            enderecamento,
                            data_chegada,
                            lote,
                            peso,
                            pei,
                            area_m2
                        ))

                # =====================================================
                # MENSAGEM / ANEXOS
                # =====================================================

                comentario = request.form.get(
                    "comentario_pedido",
                    ""
                ).strip()

                arquivos = request.files.getlist("anexos")

                tem_anexo = any(
                    f and f.filename
                    for f in arquivos
                )

                mensagem_id = None

                # cria mensagem mesmo sem comentário
                if comentario or tem_anexo:

                    cursor = conn.execute("""
                        INSERT INTO mensagens (
                            chamado_id,
                            texto,
                            autor_nome,
                            autor_id,
                            autor_tipo
                        )
                        VALUES (?, ?, ?, ?, ?)
                    """, (
                        num_pedido,
                        comentario if comentario else "",
                        usuario,
                        autor_id,
                        user_tipo
                    ))

                    mensagem_id = cursor.lastrowid

                # =====================================================
                # SALVAR ANEXOS
                # =====================================================

                if tem_anexo:

                    os.makedirs(
                        UPLOAD_FOLDER,
                        exist_ok=True
                    )

                    for f in arquivos:

                        if f and f.filename:

                            if not allowed_file(f.filename):
                                flash("Anexo ignorado por extensão não permitida.", "warning")
                                continue

                            nome_original = secure_filename(
                                f.filename
                            )
                            nome_seguro = safe_unique_filename(nome_original, "conf")

                            caminho = os.path.join(
                                UPLOAD_FOLDER,
                                nome_seguro
                            )

                            f.save(caminho)

                            conn.execute("""
                                INSERT INTO anexos (
                                    chamado_id,
                                    mensagem_id,
                                    filename,
                                    filepath
                                )
                                VALUES (?, ?, ?, ?)
                            """, (
                                num_pedido,
                                mensagem_id,
                                nome_seguro,
                                nome_seguro
                            ))

                conn.commit()

            flash(
                "Pedido atualizado com sucesso!",
                "success"
            )

        except Exception as e:

            flash(
                f"Erro ao salvar: {e}",
                "danger"
            )

        return redirect(
            url_for(
                "conferencia_itens",
                num_pedido=num_pedido
            )
        )

    # ================= BANCO LOCAL =================

    with get_db_connection() as conn:

        qtd_dict = {

            r["codpro"]: dict(r)

            for r in conn.execute("""

                SELECT
                    codpro,
                    qtd_contada,
                    tonalidade_bitola,
                    enderecamento,
                    data_chegada,
                    lote,
                    peso,
                    pei,
                    area_m2

                FROM conferencias

                WHERE numped = ?

            """, (num_pedido,))
        }

        mensagens_raw = conn.execute("""

            SELECT
                m.id,
                m.texto,
                m.autor_nome,
                m.autor_tipo,
                m.data,
                a.filename,
                a.filepath

            FROM mensagens m

            LEFT JOIN anexos a
                ON a.mensagem_id = m.id

            WHERE m.chamado_id = ?

            ORDER BY m.data ASC

        """, (num_pedido,)).fetchall()

    # ================= AGRUPAR MENSAGENS =================

    tz_brasilia = pytz.timezone("America/Sao_Paulo")

    mensagens_dict = {}

    for m in mensagens_raw:

        msg_id = m["id"]

        if msg_id not in mensagens_dict:

            data_fmt = m["data"]

            try:

                dt = datetime.strptime(
                    m["data"],
                    "%Y-%m-%d %H:%M:%S"
                )

                dt = dt.replace(tzinfo=pytz.UTC)

                data_fmt = dt.astimezone(
                    tz_brasilia
                ).strftime("%d/%m/%Y %H:%M")

            except:
                pass

            mensagens_dict[msg_id] = {
                "texto": m["texto"],
                "autor_nome": m["autor_nome"],
                "autor_tipo": m["autor_tipo"],
                "data": data_fmt,
                "anexos": []
            }

        if m["filename"]:

            mensagens_dict[msg_id]["anexos"].append({
                "filename": m["filename"],
                "filepath": m["filepath"]
            })

    mensagens = list(mensagens_dict.values())

    # ================= TOTAL PEDIDO =================

    total_pedido = sum(
        Decimal(str(i["TOTAL_PRECO_FINAL"]))
        for i in itens
        if i["TOTAL_PRECO_FINAL"]
    )

    return render_template(
        "conferencia_itens.html",
        num_pedido=num_pedido,
        itens=itens,
        mensagens=mensagens,
        qtd_dict=qtd_dict,
        user_tipo=user_tipo,
        total_pedido=total_pedido
    )

from flask import send_from_directory

@app.route("/anexos/<path:filename>")
@login_required
def ver_anexo(filename):
    if os.path.basename(filename) != filename or ".." in filename:
        abort(400)
    return send_from_directory(
        UPLOAD_FOLDER,
        filename,
        as_attachment=False
    )


# =========================================================
# HARDENING DE PRODUÇÃO: CSRF, CABEÇALHOS E ENDPOINTS AUXILIARES
# =========================================================

def csrf_token():
    token = session.get("_csrf_token")
    if not token:
        token = secrets.token_urlsafe(32)
        session["_csrf_token"] = token
    return token

@app.context_processor
def inject_security_helpers():
    return {"csrf_token": csrf_token, "static_version": "20260527-v2"}

@app.before_request
def enforce_csrf_protection():
    if request.method in ("GET", "HEAD", "OPTIONS"):
        return None
    if request.endpoint == "static":
        return None

    expected = session.get("_csrf_token")
    supplied = (
        request.form.get("_csrf_token")
        or request.headers.get("X-CSRFToken")
        or request.headers.get("X-CSRF-Token")
    )
    if not expected or not supplied or not hmac.compare_digest(str(expected), str(supplied)):
        if request.accept_mimetypes.accept_json or request.is_json:
            return jsonify({"erro": "Token CSRF inválido ou ausente."}), 400
        flash("Sessão expirada ou requisição inválida. Tente novamente.", "danger")
        return redirect(url_for("login"))
    return None

@app.after_request
def add_security_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("X-Frame-Options", "SAMEORIGIN")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "geolocation=(), microphone=(), camera=()")
    response.headers.setdefault("Cache-Control", "no-store" if session.get("user_id") else "no-cache")
    return response

@app.route("/api/conferencia/pendentes")
@login_required
def api_conferencia_pendentes():
    try:
        with get_db_connection() as conn:
            row = conn.execute(
                "SELECT COUNT(*) AS total FROM status_conferencia WHERE COALESCE(status, ?) = ?",
                (STATUS_PENDENTE, STATUS_PENDENTE),
            ).fetchone()
            total = int(row["total"] if row else 0)
        return jsonify({"total": total})
    except Exception:
        return jsonify({"total": 0})

# --- ADICIONE ESTE BLOCO NO FINAL DO SEU APP.PY ---

# Token de segurança (Adicione também nas variáveis de ambiente do Render para maior segurança)
API_SYNC_TOKEN = os.environ.get("API_SYNC_TOKEN", "ChaveSuperSecretaDoIntegrador123!")

@app.route("/api/pedidos/sincronizar", methods=["POST"])
def api_sincronizar_pedidos():
    # 1. Validação de segurança por token
    token_recebido = request.headers.get("X-Sync-Token")
    if not token_recebido or token_recebido != API_SYNC_TOKEN:
        return jsonify({"erro": "Acesso não autorizado."}), 401
    
    dados = request.json
    if not dados or "pedidos" not in dados:
        return jsonify({"erro": "Dados inválidos ou ausentes."}), 400
    
    pedidos_recebidos = dados["pedidos"]
    
    try:
        with get_db_connection() as conn:
            # 2. Garante que a tabela local exista no SQLite com as suas colunas exatas
            conn.execute("""
                CREATE TABLE IF NOT EXISTS pedidos_erp (
                    NUMPED TEXT PRIMARY KEY,
                    DATA_PEDIDO TEXT,
                    FORNECEDOR TEXT,
                    DTPREVREC TEXT
                )
            """)
            
            # 3. Insere ou atualiza os pedidos em massa
            for p in pedidos_recebidos:
                conn.execute("""
                    INSERT INTO pedidos_erp (NUMPED, DATA_PEDIDO, FORNECEDOR, DTPREVREC)
                    VALUES (?, ?, ?, ?)
                    ON CONFLICT(NUMPED) DO UPDATE SET
                        DATA_PEDIDO = excluded.DATA_PEDIDO,
                        FORNECEDOR = excluded.FORNECEDOR,
                        DTPREVREC = excluded.DTPREVREC
                """, (str(p['NUMPED']), p['DATA_PEDIDO'], p['FORNECEDOR'], p['DTPREVREC']))
            
            conn.commit()
            
        return jsonify({"status": "sucesso", "mensagem": f"{len(pedidos_recebidos)} pedidos sincronizados."}), 200
        
    except Exception as e:
        return jsonify({"erro": f"Erro ao gravar no SQLite local: {str(e)}"}), 500




